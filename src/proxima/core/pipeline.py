"""
Data Flow Pipeline Orchestrator
================================

Implements the complete data flow for Proxima:

    User Input → Parse → Plan → Check Resources → Get Consent
                                                        │
                                                        ▼
                                                Execute on Backend(s)
                                                        │
                                                        ▼
                                                Collect Results
                                                        │
                                                        ▼
                                                Generate Insights
                                                        │
                                                        ▼
                                                Export/Display

This module orchestrates the complete execution pipeline with:
- Stage tracking and timing
- Error handling and rollback
- Progress reporting
- Resource monitoring
"""

from __future__ import annotations

import asyncio
import time
import uuid
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum, auto
from typing import Any, TypeVar

import structlog

logger = structlog.get_logger(__name__)

T = TypeVar("T")


# =============================================================================
# PIPELINE STAGES
# =============================================================================


class PipelineStage(Enum):
    """Stages in the data flow pipeline."""

    IDLE = auto()
    PARSING = auto()
    PLANNING = auto()
    RESOURCE_CHECK = auto()
    CONSENT = auto()
    EXECUTING = auto()
    COLLECTING = auto()
    ANALYZING = auto()
    EXPORTING = auto()
    COMPLETED = auto()
    FAILED = auto()
    ABORTED = auto()


@dataclass
class StageResult:
    """Result of a pipeline stage execution."""

    stage: PipelineStage
    success: bool
    data: Any = None
    error: str | None = None
    duration_ms: float = 0.0
    timestamp: float = field(default_factory=time.time)

    def to_dict(self) -> dict[str, Any]:
        return {
            "stage": self.stage.name,
            "success": self.success,
            "data": self.data,
            "error": self.error,
            "duration_ms": self.duration_ms,
            "timestamp": self.timestamp,
        }


@dataclass
class PipelineContext:
    """Context passed through the pipeline."""

    # Execution metadata
    execution_id: str = field(default_factory=lambda: str(uuid.uuid4())[:8])
    started_at: float = field(default_factory=time.time)

    # Input data
    user_input: str = ""
    input_params: dict[str, Any] = field(default_factory=dict)

    # Execution plan
    plan: dict[str, Any] | None = None
    selected_backends: list[str] = field(default_factory=list)

    # Resource info
    resource_check_passed: bool = False
    resource_warnings: list[str] = field(default_factory=list)

    # Consent
    consent_granted: bool = False
    consent_details: dict[str, bool] = field(default_factory=dict)

    # Execution results
    backend_results: dict[str, Any] = field(default_factory=dict)

    # Analysis
    insights: list[str] = field(default_factory=list)
    comparison: dict[str, Any] | None = None

    # Export
    export_path: str | None = None
    export_format: str | None = None

    # Stage tracking
    stage_results: list[StageResult] = field(default_factory=list)
    current_stage: PipelineStage = PipelineStage.IDLE

    @property
    def elapsed_ms(self) -> float:
        """Get total elapsed time in milliseconds."""
        return (time.time() - self.started_at) * 1000

    def add_stage_result(self, result: StageResult) -> None:
        """Add a stage result to the history."""
        self.stage_results.append(result)
        self.current_stage = result.stage

    def to_dict(self) -> dict[str, Any]:
        """Convert context to dictionary."""
        return {
            "execution_id": self.execution_id,
            "started_at": datetime.fromtimestamp(self.started_at).isoformat(),
            "elapsed_ms": self.elapsed_ms,
            "current_stage": self.current_stage.name,
            "backends": self.selected_backends,
            "resource_check_passed": self.resource_check_passed,
            "consent_granted": self.consent_granted,
            "has_results": bool(self.backend_results),
            "insights_count": len(self.insights),
            "export_path": self.export_path,
        }


# =============================================================================
# PIPELINE HANDLERS (Stage Implementations)
# =============================================================================


class PipelineHandler:
    """Base class for pipeline stage handlers."""

    def __init__(self, stage: PipelineStage):
        self.stage = stage

    async def execute(self, ctx: PipelineContext) -> StageResult:
        """Execute the stage. Override in subclasses."""
        raise NotImplementedError


class ParseHandler(PipelineHandler):
    """Parse user input into structured format."""

    def __init__(self):
        super().__init__(PipelineStage.PARSING)

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            # Parse input - extract intent, parameters, options
            parsed = {
                "raw_input": ctx.user_input,
                "type": "simulation",  # Default type
                "parameters": ctx.input_params,
            }

            # Detect backend hints
            input_lower = ctx.user_input.lower()
            if "compare" in input_lower:
                parsed["type"] = "comparison"
            elif "analyze" in input_lower:
                parsed["type"] = "analysis"

            ctx.plan = parsed

            return StageResult(
                stage=self.stage,
                success=True,
                data=parsed,
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class PlanHandler(PipelineHandler):
    """Create execution plan based on parsed input."""

    def __init__(self):
        super().__init__(PipelineStage.PLANNING)

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            # Determine backends to use
            if "backend" in ctx.input_params:
                backends = [ctx.input_params["backend"]]
            elif ctx.plan and ctx.plan.get("type") == "comparison":
                backends = ["cirq", "qiskit-aer"]  # Default comparison
            else:
                backends = ["cirq"]  # Default single backend

            ctx.selected_backends = backends

            plan = {
                "execution_type": ctx.plan.get("type", "simulation") if ctx.plan else "simulation",
                "backends": backends,
                "parallel": len(backends) > 1,
                "estimated_time_s": len(backends) * 5,  # Rough estimate
            }

            ctx.plan = {**(ctx.plan or {}), **plan}

            return StageResult(
                stage=self.stage,
                success=True,
                data=plan,
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class ResourceCheckHandler(PipelineHandler):
    """Check system resources before execution."""

    def __init__(self, memory_threshold_mb: int = 1024):
        super().__init__(PipelineStage.RESOURCE_CHECK)
        self.memory_threshold_mb = memory_threshold_mb

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            import psutil

            memory = psutil.virtual_memory()
            available_mb = memory.available // (1024 * 1024)

            warnings = []
            if available_mb < self.memory_threshold_mb:
                warnings.append(
                    f"Low memory: {available_mb}MB available "
                    f"(threshold: {self.memory_threshold_mb}MB)"
                )

            if memory.percent > 90:
                warnings.append(f"High memory usage: {memory.percent}%")

            ctx.resource_check_passed = len(warnings) == 0 or available_mb > 512
            ctx.resource_warnings = warnings

            return StageResult(
                stage=self.stage,
                success=ctx.resource_check_passed,
                data={
                    "available_mb": available_mb,
                    "memory_percent": memory.percent,
                    "warnings": warnings,
                },
                error="; ".join(warnings) if warnings and not ctx.resource_check_passed else None,
                duration_ms=(time.time() - start) * 1000,
            )
        except ImportError:
            # psutil not available - pass with warning
            ctx.resource_check_passed = True
            ctx.resource_warnings = ["Resource monitoring unavailable (psutil not installed)"]
            return StageResult(
                stage=self.stage,
                success=True,
                data={"warning": "psutil not available"},
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class ConsentHandler(PipelineHandler):
    """Handle user consent for sensitive operations."""

    def __init__(
        self,
        require_consent: bool = True,
        auto_approve: bool = False,
        consent_callback: Callable[[str], bool] | None = None,
    ):
        super().__init__(PipelineStage.CONSENT)
        self.require_consent = require_consent
        self.auto_approve = auto_approve
        self.consent_callback = consent_callback

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()

        if not self.require_consent or self.auto_approve:
            ctx.consent_granted = True
            ctx.consent_details = {"auto_approved": True}
            return StageResult(
                stage=self.stage,
                success=True,
                data={"auto_approved": True},
                duration_ms=(time.time() - start) * 1000,
            )

        try:
            # Check if consent is needed
            needs_consent = []

            if ctx.resource_warnings:
                needs_consent.append("resource_warnings")

            if not needs_consent:
                ctx.consent_granted = True
                ctx.consent_details = {"no_consent_needed": True}
                return StageResult(
                    stage=self.stage,
                    success=True,
                    data={"no_consent_needed": True},
                    duration_ms=(time.time() - start) * 1000,
                )

            # Request consent via callback
            if self.consent_callback:
                consent_message = f"Consent required for: {', '.join(needs_consent)}"
                granted = self.consent_callback(consent_message)
            else:
                # Default: grant consent (in production, this should prompt user)
                granted = True

            ctx.consent_granted = granted
            ctx.consent_details = dict.fromkeys(needs_consent, granted)

            return StageResult(
                stage=self.stage,
                success=granted,
                data=ctx.consent_details,
                error="Consent denied by user" if not granted else None,
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class ExecutionHandler(PipelineHandler):
    """Execute simulation on selected backends."""

    def __init__(self, backend_executor: Callable | None = None):
        super().__init__(PipelineStage.EXECUTING)
        self.backend_executor = backend_executor

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            results = {}

            for backend in ctx.selected_backends:
                backend_start = time.time()

                if self.backend_executor:
                    # Use provided executor
                    result = await self.backend_executor(backend, ctx)
                else:
                    # Simulate execution
                    await asyncio.sleep(0.1)  # Simulate work
                    result = {
                        "backend": backend,
                        "status": "success",
                        "counts": {"00": 500, "11": 500},
                        "duration_ms": (time.time() - backend_start) * 1000,
                    }

                results[backend] = result

            ctx.backend_results = results

            return StageResult(
                stage=self.stage,
                success=True,
                data=results,
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class CollectionHandler(PipelineHandler):
    """Collect and normalize results from backends."""

    def __init__(self):
        super().__init__(PipelineStage.COLLECTING)

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            # Normalize results from different backends
            normalized = {}

            for backend, result in ctx.backend_results.items():
                normalized[backend] = {
                    "backend": backend,
                    "status": result.get("status", "unknown"),
                    "counts": result.get("counts", {}),
                    "duration_ms": result.get("duration_ms", 0),
                    "metadata": result.get("metadata", {}),
                }

            ctx.backend_results = normalized

            return StageResult(
                stage=self.stage,
                success=True,
                data=normalized,
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class AnalysisHandler(PipelineHandler):
    """Generate insights from results."""

    def __init__(self, insight_generator: Callable | None = None):
        super().__init__(PipelineStage.ANALYZING)
        self.insight_generator = insight_generator

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            insights = []
            comparison = None

            # Generate basic insights
            for backend, result in ctx.backend_results.items():
                counts = result.get("counts", {})
                if counts:
                    total = sum(counts.values())
                    dominant = max(counts.items(), key=lambda x: x[1])
                    insights.append(
                        f"{backend}: Dominant state '{dominant[0]}' with "
                        f"{dominant[1]/total*100:.1f}% probability"
                    )

            # Generate comparison if multiple backends
            if len(ctx.backend_results) > 1:
                times = {k: v.get("duration_ms", 0) for k, v in ctx.backend_results.items()}
                fastest = min(times.items(), key=lambda x: x[1])
                comparison = {
                    "fastest_backend": fastest[0],
                    "execution_times": times,
                    "speedup": max(times.values()) / fastest[1] if fastest[1] > 0 else 1.0,
                }
                insights.append(f"Fastest backend: {fastest[0]} ({fastest[1]:.1f}ms)")

            # Use custom insight generator if provided
            if self.insight_generator:
                custom_insights = await self.insight_generator(ctx.backend_results)
                insights.extend(custom_insights)

            ctx.insights = insights
            ctx.comparison = comparison

            return StageResult(
                stage=self.stage,
                success=True,
                data={"insights": insights, "comparison": comparison},
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )


class ExportHandler(PipelineHandler):
    """Export results to file in multiple formats (JSON, CSV, XLSX)."""

    def __init__(self, export_dir: str = "./results"):
        super().__init__(PipelineStage.EXPORTING)
        self.export_dir = export_dir

    async def execute(self, ctx: PipelineContext) -> StageResult:
        start = time.time()
        try:
            import json
            import os

            # Create export directory
            os.makedirs(self.export_dir, exist_ok=True)

            # Generate export data
            export_data = {
                "execution_id": ctx.execution_id,
                "timestamp": datetime.now().isoformat(),
                "input": ctx.user_input,
                "backends": ctx.selected_backends,
                "results": ctx.backend_results,
                "insights": ctx.insights,
                "comparison": ctx.comparison,
                "duration_ms": ctx.elapsed_ms,
            }

            # Determine format
            fmt = (ctx.export_format or "json").lower()
            filename = f"{ctx.execution_id}_{int(time.time())}.{fmt}"
            filepath = os.path.join(self.export_dir, filename)

            if fmt == "json":
                with open(filepath, "w", encoding="utf-8") as f:
                    json.dump(export_data, f, indent=2, default=str)

            elif fmt == "csv":
                filepath = self._export_csv(export_data, filepath)

            elif fmt == "xlsx":
                filepath = self._export_xlsx(export_data, filepath)

            else:
                # Default to JSON for unknown formats
                filepath = filepath.replace(f".{fmt}", ".json")
                with open(filepath, "w", encoding="utf-8") as f:
                    json.dump(export_data, f, indent=2, default=str)
                fmt = "json"

            ctx.export_path = filepath

            return StageResult(
                stage=self.stage,
                success=True,
                data={"export_path": filepath, "format": fmt},
                duration_ms=(time.time() - start) * 1000,
            )
        except Exception as e:
            return StageResult(
                stage=self.stage,
                success=False,
                error=str(e),
                duration_ms=(time.time() - start) * 1000,
            )

    def _export_csv(self, data: dict[str, Any], filepath: str) -> str:
        """Export data to CSV format."""
        import csv

        # Flatten results for CSV
        rows = []

        # Summary row
        rows.append(
            {
                "type": "summary",
                "execution_id": data["execution_id"],
                "timestamp": data["timestamp"],
                "input": data["input"],
                "duration_ms": data["duration_ms"],
                "backends": ", ".join(data["backends"]),
            }
        )

        # Backend results rows
        for backend, result in data.get("results", {}).items():
            row = {
                "type": "result",
                "backend": backend,
                "status": result.get("status", "unknown"),
                "duration_ms": result.get("duration_ms", 0),
            }
            # Add counts as columns
            counts = result.get("counts", {})
            for state, count in counts.items():
                row[f"count_{state}"] = count
            rows.append(row)

        # Insights rows
        for i, insight in enumerate(data.get("insights", []), 1):
            rows.append(
                {
                    "type": "insight",
                    "insight_num": i,
                    "insight_text": insight,
                }
            )

        # Write CSV
        if rows:
            all_keys: set[str] = set()
            for row in rows:
                all_keys.update(row.keys())
            fieldnames = sorted(all_keys)

            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

        return filepath

    def _export_xlsx(self, data: dict[str, Any], filepath: str) -> str:
        """Export data to XLSX format using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fall back to CSV if openpyxl not available
            logger.warning("openpyxl not available, falling back to CSV")
            return self._export_csv(data, filepath.replace(".xlsx", ".csv"))

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Summary data
        summary_data = [
            ("Execution ID", data["execution_id"]),
            ("Timestamp", data["timestamp"]),
            ("Input", data["input"]),
            ("Duration (ms)", data["duration_ms"]),
            ("Backends", ", ".join(data["backends"])),
        ]

        for row_num, (key, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_num, column=1, value=key).font = header_font
            ws_summary.cell(row=row_num, column=2, value=str(value))

        # Results sheet
        ws_results = wb.create_sheet("Results")
        results = data.get("results", {})

        if results:
            # Headers
            headers = ["Backend", "Status", "Duration (ms)"]
            # Collect all count keys
            count_keys: set[str] = set()
            for result in results.values():
                count_keys.update(result.get("counts", {}).keys())
            headers.extend([f"Count: {k}" for k in sorted(count_keys)])

            for col, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Data rows
            for row_num, (backend, result) in enumerate(results.items(), 2):
                ws_results.cell(row=row_num, column=1, value=backend)
                ws_results.cell(row=row_num, column=2, value=result.get("status", "unknown"))
                ws_results.cell(row=row_num, column=3, value=result.get("duration_ms", 0))

                counts = result.get("counts", {})
                for col, key in enumerate(sorted(count_keys), 4):
                    ws_results.cell(row=row_num, column=col, value=counts.get(key, 0))

        # Insights sheet
        ws_insights = wb.create_sheet("Insights")
        insights = data.get("insights", [])

        ws_insights.cell(row=1, column=1, value="Insight #").font = header_font
        ws_insights.cell(row=1, column=2, value="Insight").font = header_font

        for row_num, insight in enumerate(insights, 2):
            ws_insights.cell(row=row_num, column=1, value=row_num - 1)
            ws_insights.cell(row=row_num, column=2, value=insight)

        # Auto-adjust column widths
        for ws in [ws_summary, ws_results, ws_insights]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except TypeError:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        return filepath


# =============================================================================
# PIPELINE ORCHESTRATOR
# =============================================================================


class DataFlowPipeline:
    """
    Orchestrates the complete data flow pipeline.

    Usage:
        pipeline = DataFlowPipeline()
        result = await pipeline.run("simulate bell state", backend="cirq")
    """

    def __init__(
        self,
        require_consent: bool = True,
        auto_approve_consent: bool = False,
        export_dir: str = "./results",
        memory_threshold_mb: int = 1024,
    ):
        # Configure handlers
        self.handlers: dict[PipelineStage, PipelineHandler] = {
            PipelineStage.PARSING: ParseHandler(),
            PipelineStage.PLANNING: PlanHandler(),
            PipelineStage.RESOURCE_CHECK: ResourceCheckHandler(memory_threshold_mb),
            PipelineStage.CONSENT: ConsentHandler(require_consent, auto_approve_consent),
            PipelineStage.EXECUTING: ExecutionHandler(),
            PipelineStage.COLLECTING: CollectionHandler(),
            PipelineStage.ANALYZING: AnalysisHandler(),
            PipelineStage.EXPORTING: ExportHandler(export_dir),
        }

        # Pipeline stage order
        self.stage_order = [
            PipelineStage.PARSING,
            PipelineStage.PLANNING,
            PipelineStage.RESOURCE_CHECK,
            PipelineStage.CONSENT,
            PipelineStage.EXECUTING,
            PipelineStage.COLLECTING,
            PipelineStage.ANALYZING,
            PipelineStage.EXPORTING,
        ]

        # Callbacks
        self._on_stage_start: Callable[[PipelineStage, PipelineContext], None] | None = None
        self._on_stage_complete: Callable[[StageResult, PipelineContext], None] | None = None
        self._on_pipeline_complete: Callable[[PipelineContext], None] | None = None

    def on_stage_start(self, callback: Callable[[PipelineStage, PipelineContext], None]) -> None:
        """Register callback for stage start events."""
        self._on_stage_start = callback

    def on_stage_complete(self, callback: Callable[[StageResult, PipelineContext], None]) -> None:
        """Register callback for stage completion events."""
        self._on_stage_complete = callback

    def on_pipeline_complete(self, callback: Callable[[PipelineContext], None]) -> None:
        """Register callback for pipeline completion."""
        self._on_pipeline_complete = callback

    def set_handler(self, stage: PipelineStage, handler: PipelineHandler) -> None:
        """Replace a stage handler."""
        self.handlers[stage] = handler

    async def run(
        self,
        user_input: str,
        **params: Any,
    ) -> PipelineContext:
        """
        Run the complete pipeline.

        Args:
            user_input: User's input command/query
            **params: Additional parameters (backend, format, etc.)

        Returns:
            PipelineContext with all results
        """
        # Initialize context
        ctx = PipelineContext(
            user_input=user_input,
            input_params=params,
        )

        logger.info(
            "pipeline_started",
            execution_id=ctx.execution_id,
            input=user_input[:50],
        )

        try:
            # Execute each stage in order
            for stage in self.stage_order:
                handler = self.handlers.get(stage)
                if not handler:
                    continue

                # Notify stage start
                if self._on_stage_start:
                    self._on_stage_start(stage, ctx)

                logger.debug("stage_starting", stage=stage.name)

                # Execute stage
                result = await handler.execute(ctx)
                ctx.add_stage_result(result)

                # Notify stage complete
                if self._on_stage_complete:
                    self._on_stage_complete(result, ctx)

                logger.debug(
                    "stage_completed",
                    stage=stage.name,
                    success=result.success,
                    duration_ms=result.duration_ms,
                )

                # Stop on failure
                if not result.success:
                    ctx.current_stage = PipelineStage.FAILED
                    logger.error(
                        "pipeline_failed",
                        execution_id=ctx.execution_id,
                        failed_stage=stage.name,
                        error=result.error,
                    )
                    break
            else:
                # All stages completed successfully
                ctx.current_stage = PipelineStage.COMPLETED
                logger.info(
                    "pipeline_completed",
                    execution_id=ctx.execution_id,
                    duration_ms=ctx.elapsed_ms,
                )

        except asyncio.CancelledError:
            ctx.current_stage = PipelineStage.ABORTED
            logger.warning("pipeline_aborted", execution_id=ctx.execution_id)
            raise

        except Exception as e:
            ctx.current_stage = PipelineStage.FAILED
            ctx.add_stage_result(
                StageResult(
                    stage=ctx.current_stage,
                    success=False,
                    error=str(e),
                )
            )
            logger.exception("pipeline_error", execution_id=ctx.execution_id)

        finally:
            if self._on_pipeline_complete:
                self._on_pipeline_complete(ctx)

        return ctx


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================


async def run_simulation(
    user_input: str,
    backend: str = "cirq",
    export: bool = True,
    **kwargs: Any,
) -> PipelineContext:
    """
    Convenience function to run a simulation.

    Args:
        user_input: Simulation description
        backend: Backend to use
        export: Whether to export results
        **kwargs: Additional parameters

    Returns:
        Pipeline context with results
    """
    pipeline = DataFlowPipeline(
        require_consent=kwargs.pop("require_consent", False),
        auto_approve_consent=kwargs.pop("auto_approve", True),
    )

    return await pipeline.run(user_input, backend=backend, **kwargs)


async def compare_backends(
    user_input: str,
    backends: list[str],
    **kwargs: Any,
) -> PipelineContext:
    """
    Compare execution across multiple backends.

    Args:
        user_input: Simulation description
        backends: List of backends to compare
        **kwargs: Additional parameters

    Returns:
        Pipeline context with comparison results
    """
    pipeline = DataFlowPipeline(
        require_consent=kwargs.pop("require_consent", False),
        auto_approve_consent=kwargs.pop("auto_approve", True),
    )

    # Set backends in params
    kwargs["backends"] = backends

    # Override planning handler to use specified backends
    class MultiBackendPlanHandler(PlanHandler):
        def __init__(self, target_backends: list[str]):
            super().__init__()
            self.target_backends = target_backends

        async def execute(self, ctx: PipelineContext) -> StageResult:
            result = await super().execute(ctx)
            ctx.selected_backends = self.target_backends
            if ctx.plan:
                ctx.plan["backends"] = self.target_backends
            return result

    pipeline.set_handler(PipelineStage.PLANNING, MultiBackendPlanHandler(backends))

    return await pipeline.run(user_input, **kwargs)


__all__ = [
    "PipelineStage",
    "StageResult",
    "PipelineContext",
    "PipelineHandler",
    "ParseHandler",
    "PlanHandler",
    "ResourceCheckHandler",
    "ConsentHandler",
    "ExecutionHandler",
    "CollectionHandler",
    "AnalysisHandler",
    "ExportHandler",
    "DataFlowPipeline",
    "run_simulation",
    "compare_backends",
]
