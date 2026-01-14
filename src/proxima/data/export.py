"""Step 5.3: Export Engine - Export results in multiple formats.

Export Formats:
| Format   | Library       | Features                      |
| CSV      | csv (stdlib)  | Simple tabular data           |
| XLSX     | openpyxl      | Multiple sheets, formatting   |
| JSON     | json (stdlib) | Full data structure           |
| HTML     | jinja2        | Rich formatted reports        |
| MARKDOWN | stdlib        | Documentation-friendly format |
| YAML     | pyyaml/stdlib | Config-friendly format        |

Report Structure (XLSX):
Workbook:
- Sheet: Summary        - Overview, key metrics
- Sheet: Raw Results    - Full measurement data
- Sheet: Backend Comparison - Side-by-side metrics
- Sheet: Insights       - Generated insights
- Sheet: Metadata       - Execution details
"""

from __future__ import annotations

import csv
import io
import json
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Protocol

# Try to import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from jinja2 import BaseLoader, Environment, Template

    HAS_JINJA2 = True
except ImportError:
    HAS_JINJA2 = False

try:
    import yaml

    HAS_YAML = True
except ImportError:
    HAS_YAML = False


class ExportFormat(Enum):
    """Supported export formats."""

    CSV = "csv"
    XLSX = "xlsx"
    JSON = "json"
    HTML = "html"
    MARKDOWN = "md"
    YAML = "yaml"


@dataclass
class ExportOptions:
    """Options for export operations."""

    format: ExportFormat = ExportFormat.JSON
    output_path: Path | None = None
    include_metadata: bool = True
    include_raw_results: bool = True
    include_comparison: bool = True
    include_insights: bool = True
    pretty_print: bool = True
    timestamp_format: str = "%Y-%m-%d %H:%M:%S"
    decimal_places: int = 6
    # CSV specific
    csv_delimiter: str = ","
    csv_quoting: int = csv.QUOTE_MINIMAL
    # XLSX specific
    xlsx_freeze_panes: bool = True
    xlsx_auto_column_width: bool = True
    # HTML specific
    html_template: str | None = None
    html_inline_styles: bool = True
    # Markdown specific
    markdown_toc: bool = True
    markdown_code_blocks: bool = True
    # YAML specific
    yaml_default_flow_style: bool = False
    yaml_allow_unicode: bool = True
    # Stream export (returns string instead of writing to file)
    stream_output: bool = False


@dataclass
class ExportResult:
    """Result of an export operation."""

    success: bool
    format: ExportFormat
    output_path: Path | None
    file_size_bytes: int = 0
    export_time_ms: float = 0.0
    error: str | None = None
    warnings: list[str] = field(default_factory=list)
    # For stream exports
    content: str | bytes | None = None

    def to_dict(self) -> dict:
        return {
            "success": self.success,
            "format": self.format.value,
            "output_path": str(self.output_path) if self.output_path else None,
            "file_size_bytes": self.file_size_bytes,
            "export_time_ms": self.export_time_ms,
            "error": self.error,
            "warnings": self.warnings,
            "has_content": self.content is not None,
        }


@dataclass
class ReportData:
    """Data structure for export reports."""

    title: str = "Proxima Execution Report"
    generated_at: float = field(default_factory=time.time)

    # Summary section
    summary: dict[str, Any] = field(default_factory=dict)

    # Raw measurement results
    raw_results: list[dict[str, Any]] = field(default_factory=list)

    # Backend comparison data
    comparison: dict[str, Any] = field(default_factory=dict)

    # Generated insights
    insights: list[str] = field(default_factory=list)

    # Execution metadata
    metadata: dict[str, Any] = field(default_factory=dict)

    # Custom sections
    custom_sections: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "title": self.title,
            "generated_at": self.generated_at,
            "summary": self.summary,
            "raw_results": self.raw_results,
            "comparison": self.comparison,
            "insights": self.insights,
            "metadata": self.metadata,
            "custom_sections": self.custom_sections,
        }


class Exporter(Protocol):
    """Protocol for exporters."""

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to the specified format."""
        ...


class BaseExporter(ABC):
    """Base class for all exporters."""

    @abstractmethod
    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to the specified format."""
        ...

    def _format_timestamp(self, timestamp: float, format_str: str) -> str:
        """Format a Unix timestamp to a string."""
        return datetime.fromtimestamp(timestamp).strftime(format_str)

    def _round_floats(self, obj: Any, decimal_places: int) -> Any:
        """Recursively round floats in a data structure."""
        if isinstance(obj, float):
            return round(obj, decimal_places)
        elif isinstance(obj, dict):
            return {k: self._round_floats(v, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._round_floats(item, decimal_places) for item in obj]
        return obj

    def _flatten_dict(self, d: dict, parent_key: str = "", sep: str = ".") -> dict:
        """Flatten a nested dictionary."""
        items: list[tuple[str, Any]] = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep).items())
            else:
                items.append((new_key, v))
        return dict(items)


class JSONExporter(BaseExporter):
    """Export data to JSON format."""

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to JSON format."""
        start_time = time.perf_counter()

        try:
            # Build export data
            export_data: dict[str, Any] = {
                "title": data.title,
                "generated_at": self._format_timestamp(
                    data.generated_at, options.timestamp_format
                ),
                "summary": data.summary if data.summary else {},
            }

            if options.include_raw_results and data.raw_results:
                export_data["raw_results"] = data.raw_results

            if options.include_comparison and data.comparison:
                export_data["comparison"] = data.comparison

            if options.include_insights and data.insights:
                export_data["insights"] = data.insights

            if options.include_metadata and data.metadata:
                export_data["metadata"] = data.metadata

            if data.custom_sections:
                export_data["custom_sections"] = data.custom_sections

            # Round floats
            export_data = self._round_floats(export_data, options.decimal_places)

            # Convert to JSON
            indent = 2 if options.pretty_print else None
            json_content = json.dumps(export_data, indent=indent, ensure_ascii=False)

            # Stream output
            if options.stream_output:
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.JSON,
                    output_path=None,
                    file_size_bytes=len(json_content.encode("utf-8")),
                    export_time_ms=export_time,
                    content=json_content,
                )

            # Write to file
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.JSON,
                    output_path=None,
                    error="JSON export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(json_content, encoding="utf-8")

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.JSON,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.JSON,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )


class CSVExporter(BaseExporter):
    """Export data to CSV format."""

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to CSV format (raw results as tabular data)."""
        start_time = time.perf_counter()

        try:
            # Stream output
            if options.stream_output:
                output = io.StringIO()
                self._write_csv(data, options, output)
                csv_content = output.getvalue()
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.CSV,
                    output_path=None,
                    file_size_bytes=len(csv_content.encode("utf-8")),
                    export_time_ms=export_time,
                    content=csv_content,
                )

            # File output
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.CSV,
                    output_path=None,
                    error="CSV export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)

            with open(options.output_path, "w", newline="", encoding="utf-8") as f:
                self._write_csv(data, options, f)

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.CSV,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.CSV,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )

    def _write_csv(self, data: ReportData, options: ExportOptions, output: Any) -> None:
        """Write CSV data to a file or string buffer."""
        if not data.raw_results:
            # Write summary as key-value pairs
            writer = csv.writer(
                output, delimiter=options.csv_delimiter, quoting=options.csv_quoting
            )
            writer.writerow(["Key", "Value"])
            flat_summary = self._flatten_dict(data.summary)
            for key, value in flat_summary.items():
                if isinstance(value, float):
                    value = round(value, options.decimal_places)
                writer.writerow([key, value])
        else:
            # Write raw results as table
            if data.raw_results:
                headers = list(data.raw_results[0].keys())
                writer = csv.DictWriter(
                    output,
                    fieldnames=headers,
                    delimiter=options.csv_delimiter,
                    quoting=options.csv_quoting,
                )
                writer.writeheader()
                for row in data.raw_results:
                    # Round floats in row
                    processed_row = {}
                    for k, v in row.items():
                        if isinstance(v, float):
                            processed_row[k] = round(v, options.decimal_places)
                        else:
                            processed_row[k] = v
                    writer.writerow(processed_row)


class XLSXExporter(BaseExporter):
    """Export data to XLSX format with multiple sheets, charts, and conditional formatting."""

    # Color palette for styling
    COLORS = {
        "primary": "4472C4",
        "secondary": "5B9BD5",
        "success": "70AD47",
        "warning": "FFC000",
        "danger": "C00000",
        "header_bg": "4472C4",
        "header_fg": "FFFFFF",
        "alt_row_bg": "D6DCE4",
        "light_bg": "E7EBF0",
    }

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to XLSX format with enhanced formatting."""
        start_time = time.perf_counter()

        if not HAS_OPENPYXL:
            return ExportResult(
                success=False,
                format=ExportFormat.XLSX,
                output_path=options.output_path,
                error="openpyxl is required for XLSX export. Install with: pip install openpyxl",
            )

        try:
            wb = openpyxl.Workbook()

            # Define styles
            header_font = Font(bold=True, color=self.COLORS["header_fg"])
            header_fill = PatternFill(
                start_color=self.COLORS["header_bg"],
                end_color=self.COLORS["header_bg"],
                fill_type="solid",
            )
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            alt_row_fill = PatternFill(
                start_color=self.COLORS["alt_row_bg"],
                end_color=self.COLORS["alt_row_bg"],
                fill_type="solid",
            )
            success_fill = PatternFill(
                start_color=self.COLORS["success"],
                end_color=self.COLORS["success"],
                fill_type="solid",
            )
            warning_fill = PatternFill(
                start_color=self.COLORS["warning"],
                end_color=self.COLORS["warning"],
                fill_type="solid",
            )
            danger_fill = PatternFill(
                start_color=self.COLORS["danger"],
                end_color=self.COLORS["danger"],
                fill_type="solid",
            )

            styles = {
                "header_font": header_font,
                "header_fill": header_fill,
                "header_alignment": header_alignment,
                "thin_border": thin_border,
                "alt_row_fill": alt_row_fill,
                "success_fill": success_fill,
                "warning_fill": warning_fill,
                "danger_fill": danger_fill,
            }

            # Create Summary sheet
            ws = wb.active
            if ws is not None:
                ws.title = "Summary"
                self._write_summary_sheet(ws, data, options, styles)

            # Create Raw Results sheet
            if options.include_raw_results and data.raw_results:
                ws_raw = wb.create_sheet("Raw Results")
                self._write_raw_results_sheet(ws_raw, data, options, styles)

            # Create Backend Comparison sheet
            if options.include_comparison and data.comparison:
                ws_comp = wb.create_sheet("Backend Comparison")
                self._write_comparison_sheet(ws_comp, data, options, styles)
                # Add chart if possible
                self._add_comparison_chart(ws_comp, data)

            # Create Insights sheet
            if options.include_insights and data.insights:
                ws_insights = wb.create_sheet("Insights")
                self._write_insights_sheet(ws_insights, data, options, styles)

            # Create Metadata sheet
            if options.include_metadata and data.metadata:
                ws_meta = wb.create_sheet("Metadata")
                self._write_metadata_sheet(ws_meta, data, options, styles)

            # Stream output
            if options.stream_output:
                buffer = io.BytesIO()
                wb.save(buffer)
                xlsx_content = buffer.getvalue()
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.XLSX,
                    output_path=None,
                    file_size_bytes=len(xlsx_content),
                    export_time_ms=export_time,
                    content=xlsx_content,
                )

            # File output
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.XLSX,
                    output_path=None,
                    error="XLSX export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(options.output_path)

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.XLSX,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.XLSX,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )

    def _apply_conditional_formatting(
        self,
        ws: Any,
        cell_range: str,
        format_type: str,
    ) -> None:
        """Apply conditional formatting to a range of cells.
        
        Args:
            ws: Worksheet to apply formatting to
            cell_range: Cell range string (e.g., "B2:B10")
            format_type: Type of conditional formatting
        """
        try:
            from openpyxl.formatting.rule import (
                CellIsRule,
                ColorScaleRule,
                DataBarRule,
                FormulaRule,
            )
            
            if format_type == "data_bar":
                # Data bars for numeric values
                rule = DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=self.COLORS["primary"],
                    showValue=True,
                    minLength=None,
                    maxLength=None,
                )
                ws.conditional_formatting.add(cell_range, rule)
                
            elif format_type == "color_scale":
                # Red-Yellow-Green color scale
                rule = ColorScaleRule(
                    start_type="min",
                    start_color=self.COLORS["danger"],
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=self.COLORS["warning"],
                    end_type="max",
                    end_color=self.COLORS["success"],
                )
                ws.conditional_formatting.add(cell_range, rule)
                
            elif format_type == "status":
                # Status-based formatting (success/warning/error)
                success_rule = CellIsRule(
                    operator="equal",
                    formula=['"success"'],
                    fill=PatternFill(
                        start_color=self.COLORS["success"],
                        end_color=self.COLORS["success"],
                        fill_type="solid",
                    ),
                )
                warning_rule = CellIsRule(
                    operator="equal",
                    formula=['"warning"'],
                    fill=PatternFill(
                        start_color=self.COLORS["warning"],
                        end_color=self.COLORS["warning"],
                        fill_type="solid",
                    ),
                )
                error_rule = CellIsRule(
                    operator="equal",
                    formula=['"error"'],
                    fill=PatternFill(
                        start_color=self.COLORS["danger"],
                        end_color=self.COLORS["danger"],
                        fill_type="solid",
                    ),
                )
                ws.conditional_formatting.add(cell_range, success_rule)
                ws.conditional_formatting.add(cell_range, warning_rule)
                ws.conditional_formatting.add(cell_range, error_rule)
                
        except ImportError:
            # Conditional formatting rules not available
            pass
        except Exception:
            # Skip if conditional formatting fails
            pass

    def _add_comparison_chart(self, ws: Any, data: ReportData) -> None:
        """Add a bar chart for backend comparison data.
        
        Args:
            ws: Worksheet to add chart to
            data: Report data with comparison metrics
        """
        try:
            from openpyxl.chart import BarChart, Reference
            
            # Find execution time data if present
            execution_times = data.comparison.get("execution_times", {})
            if not execution_times or len(execution_times) < 2:
                return
            
            # Add chart data to the sheet
            chart_start_row = ws.max_row + 3
            ws.cell(row=chart_start_row, column=1, value="Backend")
            ws.cell(row=chart_start_row, column=2, value="Execution Time (ms)")
            
            for idx, (backend, time_ms) in enumerate(execution_times.items(), 1):
                ws.cell(row=chart_start_row + idx, column=1, value=backend)
                ws.cell(row=chart_start_row + idx, column=2, value=time_ms)
            
            chart_end_row = chart_start_row + len(execution_times)
            
            # Create chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Backend Execution Times"
            chart.y_axis.title = "Time (ms)"
            chart.x_axis.title = "Backend"
            
            # Data reference
            data_ref = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
            cats = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=chart_end_row)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 12
            chart.height = 8
            
            ws.add_chart(chart, "D2")
            
        except ImportError:
            # Charts not available
            pass
        except Exception:
            # Skip if chart creation fails
            pass

    def _write_summary_sheet(
        self,
        ws: Any,
        data: ReportData,
        options: ExportOptions,
        styles: dict[str, Any],
    ) -> None:
        """Write the summary sheet with enhanced formatting."""
        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        header_alignment = styles["header_alignment"]
        thin_border = styles["thin_border"]
        alt_row_fill = styles["alt_row_fill"]

        # Title row with prominent styling
        ws["A1"] = f"Report: {data.title}"
        ws["A1"].font = Font(bold=True, size=16, color=self.COLORS["primary"])
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 30

        # Generated timestamp
        ws["A2"] = "Generated:"
        ws["A2"].font = Font(bold=True)
        ws["B2"] = self._format_timestamp(data.generated_at, options.timestamp_format)
        ws.row_dimensions[2].height = 20

        # Key metrics cards (if summary has standard metrics)
        summary = data.summary
        if summary:
            # Display key metrics in a horizontal layout
            ws["A4"] = "Key Metrics"
            ws["A4"].font = Font(bold=True, size=14, color=self.COLORS["primary"])
            ws.merge_cells("A4:D4")
            
            row = 6

        # Summary table with alternating rows
        ws["A8"] = "Metric"
        ws["B8"] = "Value"
        ws["A8"].font = header_font
        ws["B8"].font = header_font
        ws["A8"].fill = header_fill
        ws["B8"].fill = header_fill
        ws["A8"].alignment = header_alignment
        ws["B8"].alignment = header_alignment
        ws["A8"].border = thin_border
        ws["B8"].border = thin_border

        row = 9
        flat_summary = self._flatten_dict(data.summary)
        for idx, (key, value) in enumerate(flat_summary.items()):
            ws[f"A{row}"] = key
            if isinstance(value, float):
                ws[f"B{row}"] = round(value, options.decimal_places)
                ws[f"B{row}"].number_format = f"0.{'0' * options.decimal_places}"
            else:
                ws[f"B{row}"] = value
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            # Alternating row colors
            if idx % 2 == 1:
                ws[f"A{row}"].fill = alt_row_fill
                ws[f"B{row}"].fill = alt_row_fill
            row += 1

        # Auto-size columns
        if options.xlsx_auto_column_width:
            self._auto_size_columns(ws)

        # Freeze panes
        if options.xlsx_freeze_panes:
            ws.freeze_panes = "A9"

    def _write_raw_results_sheet(
        self,
        ws: Any,
        data: ReportData,
        options: ExportOptions,
        styles: dict[str, Any],
    ) -> None:
        """Write the raw results sheet with conditional formatting."""
        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        header_alignment = styles["header_alignment"]
        thin_border = styles["thin_border"]
        alt_row_fill = styles["alt_row_fill"]

        if not data.raw_results:
            return

        headers = list(data.raw_results[0].keys())

        # Write headers with filter enabled
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Enable auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data.raw_results) + 1}"

        # Track numeric columns for conditional formatting
        numeric_cols = []

        # Write data with alternating row colors
        for row_idx, result in enumerate(data.raw_results, 2):
            for col_idx, header in enumerate(headers, 1):
                value = result.get(header)
                if isinstance(value, float):
                    value = round(value, options.decimal_places)
                    if col_idx not in numeric_cols:
                        numeric_cols.append(col_idx)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill
                # Number formatting for floats
                if isinstance(value, float):
                    cell.number_format = f"0.{'0' * options.decimal_places}"

        # Apply data bars to numeric columns
        for col_idx in numeric_cols:
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{len(data.raw_results) + 1}"
            self._apply_conditional_formatting(ws, cell_range, "data_bar")

        # Auto-size columns
        if options.xlsx_auto_column_width:
            self._auto_size_columns(ws)

        # Freeze panes
        if options.xlsx_freeze_panes:
            ws.freeze_panes = "A2"

    def _write_comparison_sheet(
        self,
        ws: Any,
        data: ReportData,
        options: ExportOptions,
        styles: dict[str, Any],
    ) -> None:
        """Write the backend comparison sheet with highlighting."""
        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        header_alignment = styles["header_alignment"]
        thin_border = styles["thin_border"]
        alt_row_fill = styles["alt_row_fill"]
        success_fill = styles["success_fill"]

        ws["A1"] = "Key"
        ws["B1"] = "Value"
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill
        ws["A1"].alignment = header_alignment
        ws["B1"].alignment = header_alignment
        ws["A1"].border = thin_border
        ws["B1"].border = thin_border

        row = 2
        flat_comparison = self._flatten_dict(data.comparison)
        for idx, (key, value) in enumerate(flat_comparison.items()):
            ws[f"A{row}"] = key
            if isinstance(value, float):
                ws[f"B{row}"] = round(value, options.decimal_places)
                ws[f"B{row}"].number_format = f"0.{'0' * options.decimal_places}"
            else:
                ws[f"B{row}"] = value
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            
            # Highlight important metrics
            if "fastest" in key.lower() or "best" in key.lower() or "recommended" in key.lower():
                ws[f"A{row}"].fill = success_fill
                ws[f"B{row}"].fill = success_fill
                ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
                ws[f"B{row}"].font = Font(bold=True, color="FFFFFF")
            elif idx % 2 == 1:
                ws[f"A{row}"].fill = alt_row_fill
                ws[f"B{row}"].fill = alt_row_fill
                
            row += 1

        if options.xlsx_auto_column_width:
            self._auto_size_columns(ws)

        if options.xlsx_freeze_panes:
            ws.freeze_panes = "A2"

    def _write_insights_sheet(
        self,
        ws: Any,
        data: ReportData,
        options: ExportOptions,
        styles: dict[str, Any],
    ) -> None:
        """Write the insights sheet with visual formatting."""
        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        header_alignment = styles["header_alignment"]
        thin_border = styles["thin_border"]

        ws["A1"] = "#"
        ws["B1"] = "Insight"
        ws["C1"] = "Category"
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["C1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill
        ws["C1"].fill = header_fill
        ws["A1"].alignment = header_alignment
        ws["B1"].alignment = header_alignment
        ws["C1"].alignment = header_alignment

        for row, insight in enumerate(data.insights, 2):
            ws[f"A{row}"] = row - 1
            ws[f"B{row}"] = insight
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            ws[f"C{row}"].border = thin_border
            
            # Categorize insights
            insight_lower = insight.lower()
            if "fastest" in insight_lower or "best" in insight_lower:
                ws[f"C{row}"] = "Performance"
                ws[f"A{row}"].fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                ws[f"B{row}"].fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                ws[f"C{row}"].fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            elif "warning" in insight_lower or "note" in insight_lower:
                ws[f"C{row}"] = "Warning"
                ws[f"A{row}"].fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                ws[f"B{row}"].fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                ws[f"C{row}"].fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            else:
                ws[f"C{row}"] = "Info"
            
            # Text wrapping for long insights
            ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(15, min(60, len(insight) // 3))

        if options.xlsx_auto_column_width:
            self._auto_size_columns(ws)
            # Set insight column to a reasonable width
            ws.column_dimensions["B"].width = 60

        if options.xlsx_freeze_panes:
            ws.freeze_panes = "A2"

    def _write_metadata_sheet(
        self,
        ws: Any,
        data: ReportData,
        options: ExportOptions,
        styles: dict[str, Any],
    ) -> None:
        """Write the metadata sheet."""
        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        header_alignment = styles["header_alignment"]
        thin_border = styles["thin_border"]
        alt_row_fill = styles["alt_row_fill"]

        ws["A1"] = "Key"
        ws["B1"] = "Value"
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill
        ws["A1"].alignment = header_alignment
        ws["B1"].alignment = header_alignment
        ws["A1"].border = thin_border
        ws["B1"].border = thin_border

        row = 2
        flat_metadata = self._flatten_dict(data.metadata)
        for idx, (key, value) in enumerate(flat_metadata.items()):
            ws[f"A{row}"] = key
            if isinstance(value, float):
                ws[f"B{row}"] = round(value, options.decimal_places)
            else:
                ws[f"B{row}"] = value
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            if idx % 2 == 1:
                ws[f"A{row}"].fill = alt_row_fill
                ws[f"B{row}"].fill = alt_row_fill
            row += 1

        if options.xlsx_auto_column_width:
            self._auto_size_columns(ws)

        if options.xlsx_freeze_panes:
            ws.freeze_panes = "A2"

    def _auto_size_columns(self, ws: Any) -> None:
        """Auto-size columns based on content."""
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                # Skip merged cells by checking for value attribute
                if hasattr(cell, "value") and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass
            if max_length > 0:
                column_letter = get_column_letter(col_idx)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width


class HTMLExporter(BaseExporter):
    """Export data to HTML format with rich Jinja2 templates."""

    # Path to bundled templates
    TEMPLATE_DIR = Path(__file__).parent / "templates"

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to HTML format."""
        start_time = time.perf_counter()

        if HAS_JINJA2:
            return self._export_with_jinja2(data, options, start_time)
        else:
            return self._export_basic_html(data, options, start_time)

    def _load_template(self, options: ExportOptions) -> str:
        """Load the HTML template from file or options."""
        # Use custom template if provided
        if options.html_template:
            return options.html_template
        
        # Try to load bundled template
        template_path = self.TEMPLATE_DIR / "report.html.j2"
        if template_path.exists():
            return template_path.read_text(encoding="utf-8")
        
        # Fallback to embedded default template
        return self._get_default_template()

    def _get_default_template(self) -> str:
        """Return the default embedded HTML template."""
        return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    {% if inline_styles %}
    <style>
        :root {
            --primary-color: #4472C4;
            --primary-light: #5B9BD5;
            --success-color: #28a745;
            --warning-color: #ffc107;
            --danger-color: #dc3545;
            --text-color: #333;
            --text-muted: #6c757d;
            --bg-color: #f8f9fa;
            --border-color: #dee2e6;
        }
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            color: var(--text-color);
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 2rem;
        }
        .container { max-width: 1200px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1), 0 8px 32px rgba(0,0,0,0.15); padding: 2.5rem; }
        .header { text-align: center; margin-bottom: 2rem; padding-bottom: 1.5rem; border-bottom: 3px solid var(--primary-color); }
        .header h1 { color: var(--primary-color); font-size: 2.5rem; margin-bottom: 0.5rem; }
        .timestamp { color: var(--text-muted); font-size: 0.9rem; margin-top: 0.5rem; }
        h2 { color: var(--primary-color); margin: 2rem 0 1rem; border-bottom: 2px solid var(--primary-light); padding-bottom: 0.5rem; }
        .metrics-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1rem; margin: 1.5rem 0; }
        .metric-card { background: linear-gradient(135deg, var(--primary-color), var(--primary-light)); color: white; padding: 1.25rem; border-radius: 10px; text-align: center; }
        .metric-card.success { background: linear-gradient(135deg, var(--success-color), #38d96a); }
        .metric-value { font-size: 2rem; font-weight: 700; }
        .metric-label { font-size: 0.85rem; opacity: 0.9; }
        table { width: 100%; border-collapse: collapse; margin: 1rem 0; border-radius: 8px; overflow: hidden; }
        th { background: var(--primary-color); color: white; padding: 14px 16px; text-align: left; }
        td { padding: 12px 16px; border-bottom: 1px solid var(--border-color); }
        tr:nth-child(even) { background: #f8f9fa; }
        tr:hover { background: #e9ecef; }
        .insight { background: linear-gradient(135deg, #e7f3ff, #f0f7ff); border-left: 4px solid var(--primary-color); padding: 1rem; margin: 0.75rem 0; border-radius: 0 8px 8px 0; }
        .insight.performance { background: linear-gradient(135deg, #d4edda, #e8f5e9); border-left-color: var(--success-color); }
        .insight.warning { background: linear-gradient(135deg, #fff3cd, #fffde7); border-left-color: var(--warning-color); }
        .footer { margin-top: 3rem; padding-top: 1.5rem; border-top: 1px solid var(--border-color); text-align: center; color: var(--text-muted); font-size: 0.85rem; }
    </style>
    {% endif %}
</head>
<body>
    <div class="container">
        <header class="header">
            <h1>{{ title }}</h1>
            <p class="timestamp">Generated: {{ generated_at }}</p>
        </header>

        {% if key_metrics %}
        <section>
            <h2>Key Metrics</h2>
            <div class="metrics-grid">
                {% for metric in key_metrics %}
                <div class="metric-card {{ metric.type|default('') }}">
                    <div class="metric-value">{{ metric.value }}</div>
                    <div class="metric-label">{{ metric.label }}</div>
                </div>
                {% endfor %}
            </div>
        </section>
        {% endif %}

        {% if summary %}
        <section>
            <h2>Summary</h2>
            <table><tr><th>Metric</th><th>Value</th></tr>
            {% for key, value in summary.items() %}<tr><td>{{ key }}</td><td>{{ value }}</td></tr>{% endfor %}
            </table>
        </section>
        {% endif %}

        {% if raw_results %}
        <section>
            <h2>Raw Results</h2>
            <table><tr>{% for header in raw_results[0].keys() %}<th>{{ header }}</th>{% endfor %}</tr>
            {% for row in raw_results %}<tr>{% for value in row.values() %}<td>{{ value }}</td>{% endfor %}</tr>{% endfor %}
            </table>
        </section>
        {% endif %}

        {% if comparison %}
        <section>
            <h2>Backend Comparison</h2>
            <table><tr><th>Key</th><th>Value</th></tr>
            {% for key, value in comparison.items() %}<tr><td>{{ key }}</td><td>{{ value }}</td></tr>{% endfor %}
            </table>
        </section>
        {% endif %}

        {% if insights %}
        <section>
            <h2>Insights</h2>
            {% for insight in insights %}
            <div class="insight {% if 'fastest' in insight|lower %}performance{% elif 'warning' in insight|lower %}warning{% endif %}">{{ insight }}</div>
            {% endfor %}
        </section>
        {% endif %}

        {% if metadata %}
        <section>
            <h2>Metadata</h2>
            <table><tr><th>Key</th><th>Value</th></tr>
            {% for key, value in metadata.items() %}<tr><td>{{ key }}</td><td>{{ value }}</td></tr>{% endfor %}
            </table>
        </section>
        {% endif %}

        <footer class="footer">
            <p>Generated by <strong>Proxima</strong> - Quantum Simulation Framework</p>
        </footer>
    </div>
</body>
</html>
"""

    def _extract_key_metrics(self, data: ReportData) -> list[dict]:
        """Extract key metrics for card display."""
        metrics = []
        summary = data.summary or {}
        comparison = data.comparison or {}
        
        # Look for common metrics
        metric_keys = [
            ("total_shots", "Total Shots", ""),
            ("execution_time_ms", "Execution Time", "success"),
            ("backends_compared", "Backends", ""),
            ("success_rate", "Success Rate", "success"),
            ("fastest_backend", "Fastest", "success"),
            ("recommended_backend", "Recommended", "success"),
        ]
        
        for key, label, card_type in metric_keys:
            if key in summary:
                metrics.append({
                    "label": label,
                    "value": summary[key],
                    "type": card_type,
                })
            elif key in comparison:
                metrics.append({
                    "label": label,
                    "value": comparison[key],
                    "type": card_type,
                })
        
        return metrics[:6]  # Limit to 6 metrics

    def _export_with_jinja2(
        self, data: ReportData, options: ExportOptions, start_time: float
    ) -> ExportResult:
        """Export using Jinja2 templates."""
        try:
            template_str = self._load_template(options)
            template = Template(template_str)

            # Flatten nested dicts for display
            flat_summary = self._flatten_dict(data.summary) if data.summary else {}
            flat_comparison = (
                self._flatten_dict(data.comparison) if data.comparison else {}
            )
            flat_metadata = self._flatten_dict(data.metadata) if data.metadata else {}

            # Round floats
            flat_summary = self._round_floats(flat_summary, options.decimal_places)
            flat_comparison = self._round_floats(
                flat_comparison, options.decimal_places
            )
            flat_metadata = self._round_floats(flat_metadata, options.decimal_places)

            # Extract key metrics for cards
            key_metrics = self._extract_key_metrics(data)

            html = template.render(
                title=data.title,
                description=data.metadata.get("description", "") if data.metadata else "",
                generated_at=self._format_timestamp(
                    data.generated_at, options.timestamp_format
                ),
                key_metrics=key_metrics,
                summary=flat_summary if data.summary else None,
                raw_results=data.raw_results if options.include_raw_results else [],
                comparison=flat_comparison if options.include_comparison else None,
                insights=data.insights if options.include_insights else [],
                metadata=flat_metadata if options.include_metadata else None,
                custom_sections=getattr(data, "custom_sections", {}),
                inline_styles=options.html_inline_styles,
                show_toc=True,
                report_id=getattr(data, "report_id", None),
            )

            # Stream output
            if options.stream_output:
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.HTML,
                    output_path=None,
                    file_size_bytes=len(html.encode("utf-8")),
                    export_time_ms=export_time,
                    content=html,
                )

            # Write to file
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.HTML,
                    output_path=None,
                    error="HTML export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(html, encoding="utf-8")

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.HTML,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.HTML,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )

    def _export_basic_html(
        self, data: ReportData, options: ExportOptions, start_time: float
    ) -> ExportResult:
        """Fallback HTML export without Jinja2."""
        try:
            # Build simple HTML
            html_parts = [
                "<!DOCTYPE html>",
                "<html><head><meta charset='UTF-8'>",
                f"<title>{data.title}</title>",
                "<style>body{font-family:sans-serif;margin:40px}table{border-collapse:collapse}th,td{border:1px solid #ddd;padding:8px}th{background:#4472C4;color:white}</style>",
                "</head><body>",
                f"<h1>{data.title}</h1>",
                f"<p>Generated: {self._format_timestamp(data.generated_at, options.timestamp_format)}</p>",
            ]

            # Summary
            if data.summary:
                html_parts.append(
                    "<h2>Summary</h2><table><tr><th>Metric</th><th>Value</th></tr>"
                )
                for k, v in self._flatten_dict(data.summary).items():
                    html_parts.append(f"<tr><td>{k}</td><td>{v}</td></tr>")
                html_parts.append("</table>")

            # Insights
            if data.insights:
                html_parts.append("<h2>Insights</h2><ul>")
                for insight in data.insights:
                    html_parts.append(f"<li>{insight}</li>")
                html_parts.append("</ul>")

            html_parts.append("</body></html>")

            html = "\n".join(html_parts)

            # Stream output
            if options.stream_output:
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.HTML,
                    output_path=None,
                    file_size_bytes=len(html.encode("utf-8")),
                    export_time_ms=export_time,
                    content=html,
                    warnings=["Jinja2 not installed, using basic HTML template"],
                )

            # File output
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.HTML,
                    output_path=None,
                    error="HTML export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(html, encoding="utf-8")

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.HTML,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
                warnings=["Jinja2 not installed, using basic HTML template"],
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.HTML,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )


class MarkdownExporter(BaseExporter):
    """Export data to Markdown format with Jinja2 template support."""

    # Path to bundled templates
    TEMPLATE_DIR = Path(__file__).parent / "templates"

    def _load_template(self, options: ExportOptions) -> str:
        """Load the Markdown template from file or default."""
        # Try to load bundled template
        template_path = self.TEMPLATE_DIR / "report.md.j2"
        if template_path.exists():
            return template_path.read_text(encoding="utf-8")
        
        # Fallback to embedded default template
        return self._get_default_template()

    def _get_default_template(self) -> str:
        """Return the default embedded Markdown template."""
        return """# {{ title }}

*Generated: {{ generated_at }}*

{% if show_toc %}
## Table of Contents

{% if summary %}- [Summary](#summary)
{% endif %}{% if raw_results %}- [Raw Results](#raw-results)
{% endif %}{% if comparison %}- [Backend Comparison](#backend-comparison)
{% endif %}{% if insights %}- [Insights](#insights)
{% endif %}{% if metadata %}- [Metadata](#metadata)
{% endif %}
{% endif %}

{% if summary %}
## Summary

| Metric | Value |
|--------|-------|
{% for key, value in summary.items() %}| {{ key }} | {{ value }} |
{% endfor %}
{% endif %}

{% if raw_results %}
## Raw Results

| {% for header in raw_results[0].keys() %}{{ header }} | {% endfor %}

|{% for header in raw_results[0].keys() %}---|{% endfor %}

{% for row in raw_results %}| {% for value in row.values() %}{{ value }} | {% endfor %}
{% endfor %}
{% endif %}

{% if comparison %}
## Backend Comparison

| Key | Value |
|-----|-------|
{% for key, value in comparison.items() %}| {{ key }} | {{ value }} |
{% endfor %}
{% endif %}

{% if insights %}
## Insights

{% for insight in insights %}{% if 'fastest' in insight|lower or 'success' in insight|lower %}✅ {{ insight }}
{% elif 'warning' in insight|lower or 'slow' in insight|lower %}⚠️ {{ insight }}
{% else %}ℹ️ {{ insight }}
{% endif %}{% endfor %}
{% endif %}

{% if metadata %}
## Metadata

```
{% for key, value in metadata.items() %}{{ key }}: {{ value }}
{% endfor %}```
{% endif %}

---
*Generated by **Proxima** - Quantum Simulation Framework*
"""

    def _export_with_jinja2(self, data: ReportData, options: ExportOptions) -> str:
        """Export using Jinja2 templates."""
        template_str = self._load_template(options)
        template = Template(template_str)

        # Flatten nested dicts for display
        flat_summary = self._flatten_dict(data.summary) if data.summary else {}
        flat_comparison = self._flatten_dict(data.comparison) if data.comparison else {}
        flat_metadata = self._flatten_dict(data.metadata) if data.metadata else {}

        # Round floats
        flat_summary = self._round_floats(flat_summary, options.decimal_places)
        flat_comparison = self._round_floats(flat_comparison, options.decimal_places)
        flat_metadata = self._round_floats(flat_metadata, options.decimal_places)

        return template.render(
            title=data.title,
            description=data.metadata.get("description", "") if data.metadata else "",
            generated_at=self._format_timestamp(data.generated_at, options.timestamp_format),
            summary=flat_summary if data.summary else None,
            raw_results=data.raw_results if options.include_raw_results else [],
            comparison=flat_comparison if options.include_comparison else None,
            insights=data.insights if options.include_insights else [],
            metadata=flat_metadata if options.include_metadata else None,
            custom_sections=getattr(data, "custom_sections", {}),
            show_toc=options.markdown_toc,
            use_code_blocks=options.markdown_code_blocks,
        )

    def _export_basic(self, data: ReportData, options: ExportOptions) -> str:
        """Export without Jinja2 (basic fallback)."""
        md_parts: list[str] = []

        # Title
        md_parts.append(f"# {data.title}")
        md_parts.append("")
        md_parts.append(
            f"*Generated: {self._format_timestamp(data.generated_at, options.timestamp_format)}*"
        )
        md_parts.append("")

        # Table of contents
        if options.markdown_toc:
            md_parts.append("## Table of Contents")
            md_parts.append("")
            if data.summary:
                md_parts.append("- [Summary](#summary)")
            if options.include_raw_results and data.raw_results:
                md_parts.append("- [Raw Results](#raw-results)")
            if options.include_comparison and data.comparison:
                md_parts.append("- [Backend Comparison](#backend-comparison)")
            if options.include_insights and data.insights:
                md_parts.append("- [Insights](#insights)")
            if options.include_metadata and data.metadata:
                md_parts.append("- [Metadata](#metadata)")
            md_parts.append("")

        # Summary section
        if data.summary:
            md_parts.append("## Summary")
            md_parts.append("")
            md_parts.append("| Metric | Value |")
            md_parts.append("|--------|-------|")
            flat_summary = self._flatten_dict(data.summary)
            for key, value in flat_summary.items():
                if isinstance(value, float):
                    value = round(value, options.decimal_places)
                md_parts.append(f"| {key} | {value} |")
            md_parts.append("")

        # Raw Results section
        if options.include_raw_results and data.raw_results:
            md_parts.append("## Raw Results")
            md_parts.append("")
            headers = list(data.raw_results[0].keys())
            md_parts.append("| " + " | ".join(headers) + " |")
            md_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
            for row in data.raw_results:
                values = []
                for h in headers:
                    v = row.get(h, "")
                    if isinstance(v, float):
                        v = round(v, options.decimal_places)
                    values.append(str(v) if v is not None else "N/A")
                md_parts.append("| " + " | ".join(values) + " |")
            md_parts.append("")

        # Comparison section
        if options.include_comparison and data.comparison:
            md_parts.append("## Backend Comparison")
            md_parts.append("")
            md_parts.append("| Key | Value |")
            md_parts.append("|-----|-------|")
            flat_comparison = self._flatten_dict(data.comparison)
            for key, value in flat_comparison.items():
                if isinstance(value, float):
                    value = round(value, options.decimal_places)
                md_parts.append(f"| {key} | {value} |")
            md_parts.append("")

        # Insights section
        if options.include_insights and data.insights:
            md_parts.append("## Insights")
            md_parts.append("")
            for insight in data.insights:
                md_parts.append(f"- {insight}")
            md_parts.append("")

        # Metadata section
        if options.include_metadata and data.metadata:
            md_parts.append("## Metadata")
            md_parts.append("")
            if options.markdown_code_blocks:
                md_parts.append("```")
                flat_metadata = self._flatten_dict(data.metadata)
                for key, value in flat_metadata.items():
                    md_parts.append(f"{key}: {value}")
                md_parts.append("```")
            else:
                md_parts.append("| Key | Value |")
                md_parts.append("|-----|-------|")
                flat_metadata = self._flatten_dict(data.metadata)
                for key, value in flat_metadata.items():
                    md_parts.append(f"| {key} | {value} |")
            md_parts.append("")

        # Custom sections
        if data.custom_sections:
            for section_name, section_data in data.custom_sections.items():
                md_parts.append(f"## {section_name}")
                md_parts.append("")
                if isinstance(section_data, dict):
                    md_parts.append("| Key | Value |")
                    md_parts.append("|-----|-------|")
                    flat_section = self._flatten_dict(section_data)
                    for key, value in flat_section.items():
                        md_parts.append(f"| {key} | {value} |")
                elif isinstance(section_data, list):
                    for item in section_data:
                        md_parts.append(f"- {item}")
                else:
                    md_parts.append(str(section_data))
                md_parts.append("")

        md_parts.append("---")
        md_parts.append("*Generated by **Proxima** - Quantum Simulation Framework*")

        return "\n".join(md_parts)

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to Markdown format."""
        start_time = time.perf_counter()

        try:
            # Use Jinja2 if available
            if HAS_JINJA2:
                md_content = self._export_with_jinja2(data, options)
            else:
                md_content = self._export_basic(data, options)

            # Stream output
            if options.stream_output:
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.MARKDOWN,
                    output_path=None,
                    file_size_bytes=len(md_content.encode("utf-8")),
                    export_time_ms=export_time,
                    content=md_content,
                )

            # File output
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.MARKDOWN,
                    output_path=None,
                    error="Markdown export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(md_content, encoding="utf-8")

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.MARKDOWN,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.MARKDOWN,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )


class YAMLExporter(BaseExporter):
    """Export data to YAML format."""

    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to YAML format."""
        start_time = time.perf_counter()

        try:
            # Build export data
            export_data: dict[str, Any] = {
                "title": data.title,
                "generated_at": self._format_timestamp(
                    data.generated_at, options.timestamp_format
                ),
                "summary": data.summary if data.summary else {},
            }

            if options.include_raw_results and data.raw_results:
                export_data["raw_results"] = data.raw_results

            if options.include_comparison and data.comparison:
                export_data["comparison"] = data.comparison

            if options.include_insights and data.insights:
                export_data["insights"] = data.insights

            if options.include_metadata and data.metadata:
                export_data["metadata"] = data.metadata

            if data.custom_sections:
                export_data["custom_sections"] = data.custom_sections

            # Round floats
            export_data = self._round_floats(export_data, options.decimal_places)

            # Convert to YAML
            if HAS_YAML:
                yaml_content = yaml.dump(
                    export_data,
                    default_flow_style=options.yaml_default_flow_style,
                    allow_unicode=options.yaml_allow_unicode,
                    sort_keys=False,
                )
            else:
                # Fallback: simple YAML-like format without pyyaml
                yaml_content = self._simple_yaml_dump(export_data)

            # Stream output
            if options.stream_output:
                export_time = (time.perf_counter() - start_time) * 1000
                return ExportResult(
                    success=True,
                    format=ExportFormat.YAML,
                    output_path=None,
                    file_size_bytes=len(yaml_content.encode("utf-8")),
                    export_time_ms=export_time,
                    content=yaml_content,
                    warnings=(
                        []
                        if HAS_YAML
                        else ["PyYAML not installed, using simple YAML format"]
                    ),
                )

            # File output
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.YAML,
                    output_path=None,
                    error="YAML export requires an output path or stream_output=True",
                )

            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(yaml_content, encoding="utf-8")

            file_size = options.output_path.stat().st_size
            export_time = (time.perf_counter() - start_time) * 1000

            return ExportResult(
                success=True,
                format=ExportFormat.YAML,
                output_path=options.output_path,
                file_size_bytes=file_size,
                export_time_ms=export_time,
                warnings=(
                    []
                    if HAS_YAML
                    else ["PyYAML not installed, using simple YAML format"]
                ),
            )

        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.YAML,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )

    def _simple_yaml_dump(self, data: Any, indent: int = 0) -> str:
        """Simple YAML serialization without pyyaml."""
        lines: list[str] = []
        prefix = "  " * indent

        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, (dict, list)) and value:
                    lines.append(f"{prefix}{key}:")
                    lines.append(self._simple_yaml_dump(value, indent + 1))
                elif isinstance(value, str):
                    # Quote strings that might need it
                    if ":" in value or "\n" in value or value.startswith("-"):
                        lines.append(f'{prefix}{key}: "{value}"')
                    else:
                        lines.append(f"{prefix}{key}: {value}")
                elif value is None:
                    lines.append(f"{prefix}{key}: null")
                elif isinstance(value, bool):
                    lines.append(f"{prefix}{key}: {str(value).lower()}")
                else:
                    lines.append(f"{prefix}{key}: {value}")
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, (dict, list)):
                    lines.append(f"{prefix}-")
                    lines.append(self._simple_yaml_dump(item, indent + 1))
                elif isinstance(item, str):
                    if ":" in item or "\n" in item:
                        lines.append(f'{prefix}- "{item}"')
                    else:
                        lines.append(f"{prefix}- {item}")
                else:
                    lines.append(f"{prefix}- {item}")
        else:
            lines.append(f"{prefix}{data}")

        return "\n".join(lines)


class ExportEngine:
    """Main export engine that handles all export formats.

    Export Formats:
    - CSV:      Simple tabular data (csv stdlib)
    - XLSX:     Multiple sheets, formatting (openpyxl)
    - JSON:     Full data structure (json stdlib)
    - HTML:     Rich formatted reports (jinja2)
    - MARKDOWN: Documentation-friendly format (stdlib)
    - YAML:     Config-friendly format (pyyaml/stdlib)
    """

    def __init__(self) -> None:
        self._exporters: dict[ExportFormat, BaseExporter] = {
            ExportFormat.JSON: JSONExporter(),
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.XLSX: XLSXExporter(),
            ExportFormat.HTML: HTMLExporter(),
            ExportFormat.MARKDOWN: MarkdownExporter(),
            ExportFormat.YAML: YAMLExporter(),
        }

    def register_exporter(self, format: ExportFormat, exporter: BaseExporter) -> None:
        """Register a custom exporter for a format."""
        self._exporters[format] = exporter

    def get_available_formats(self) -> list[ExportFormat]:
        """Get list of available export formats."""
        available = [
            ExportFormat.JSON,
            ExportFormat.CSV,
            ExportFormat.MARKDOWN,
            ExportFormat.YAML,
        ]  # Always available
        if HAS_OPENPYXL:
            available.append(ExportFormat.XLSX)
        available.append(ExportFormat.HTML)  # Always available (with fallback)
        return available

    def export(
        self,
        data: ReportData,
        format: ExportFormat = ExportFormat.JSON,
        output_path: Path | None = None,
        **kwargs: Any,
    ) -> ExportResult:
        """Export data to the specified format.

        Args:
            data: ReportData to export
            format: Export format (JSON, CSV, XLSX, HTML, MARKDOWN, YAML)
            output_path: Output file path
            **kwargs: Additional options passed to ExportOptions

        Returns:
            ExportResult with export status
        """
        options = ExportOptions(
            format=format,
            output_path=output_path,
            **kwargs,
        )

        exporter = self._exporters.get(format)
        if not exporter:
            return ExportResult(
                success=False,
                format=format,
                output_path=output_path,
                error=f"No exporter registered for format: {format.value}",
            )

        return exporter.export(data, options)

    def export_to_string(
        self,
        data: ReportData,
        format: ExportFormat = ExportFormat.JSON,
        **kwargs: Any,
    ) -> ExportResult:
        """Export data to a string (stream export).

        Args:
            data: ReportData to export
            format: Export format (JSON, CSV, HTML, MARKDOWN, YAML)
            **kwargs: Additional options passed to ExportOptions

        Returns:
            ExportResult with content field containing the exported string
        """
        return self.export(data, format, stream_output=True, **kwargs)

    def export_to_bytes(
        self,
        data: ReportData,
        format: ExportFormat = ExportFormat.XLSX,
        **kwargs: Any,
    ) -> ExportResult:
        """Export data to bytes (stream export for binary formats).

        Args:
            data: ReportData to export
            format: Export format (XLSX)
            **kwargs: Additional options passed to ExportOptions

        Returns:
            ExportResult with content field containing the exported bytes
        """
        return self.export(data, format, stream_output=True, **kwargs)

    def export_all(
        self,
        data: ReportData,
        output_dir: Path,
        base_name: str = "report",
        formats: list[ExportFormat] | None = None,
    ) -> dict[ExportFormat, ExportResult]:
        """Export data to multiple formats.

        Args:
            data: ReportData to export
            output_dir: Directory for output files
            base_name: Base filename (without extension)
            formats: List of formats to export (default: all available)

        Returns:
            Dict mapping format to ExportResult
        """
        if formats is None:
            formats = self.get_available_formats()

        results: dict[ExportFormat, ExportResult] = {}

        for format in formats:
            extension = format.value
            output_path = output_dir / f"{base_name}.{extension}"
            results[format] = self.export(data, format, output_path)

        return results


# Convenience functions
def export_to_json(
    data: ReportData,
    output_path: Path,
    pretty: bool = True,
) -> ExportResult:
    """Export data to JSON format."""
    engine = ExportEngine()
    return engine.export(data, ExportFormat.JSON, output_path, pretty_print=pretty)


def export_to_csv(
    data: ReportData,
    output_path: Path,
) -> ExportResult:
    """Export data to CSV format."""
    engine = ExportEngine()
    return engine.export(data, ExportFormat.CSV, output_path)


def export_to_xlsx(
    data: ReportData,
    output_path: Path,
) -> ExportResult:
    """Export data to XLSX format."""
    engine = ExportEngine()
    return engine.export(data, ExportFormat.XLSX, output_path)


def export_to_html(
    data: ReportData,
    output_path: Path,
) -> ExportResult:
    """Export data to HTML format."""
    engine = ExportEngine()
    return engine.export(data, ExportFormat.HTML, output_path)


def export_to_markdown(
    data: ReportData,
    output_path: Path,
    toc: bool = True,
) -> ExportResult:
    """Export data to Markdown format."""
    engine = ExportEngine()
    return engine.export(data, ExportFormat.MARKDOWN, output_path, markdown_toc=toc)


def export_to_yaml(
    data: ReportData,
    output_path: Path,
) -> ExportResult:
    """Export data to YAML format."""
    engine = ExportEngine()
    return engine.export(data, ExportFormat.YAML, output_path)


def export_to_string(
    data: ReportData,
    format: ExportFormat = ExportFormat.JSON,
    **kwargs: Any,
) -> str | None:
    """Export data to a string.

    Returns:
        The exported string, or None if export failed.
    """
    engine = ExportEngine()
    result = engine.export_to_string(data, format, **kwargs)
    if result.success and result.content:
        return (
            result.content
            if isinstance(result.content, str)
            else result.content.decode("utf-8")
        )
    return None


# =============================================================================
# Advanced Export Templates (Feature - Export Engine)
# =============================================================================


@dataclass
class TemplateVariable:
    """Variable definition for export templates."""
    
    name: str
    type: str  # 'string', 'number', 'list', 'dict', 'date'
    description: str = ""
    default: Any = None
    required: bool = False


@dataclass 
class ExportTemplate:
    """Advanced export template definition.
    
    Supports:
    - Custom Jinja2 templates
    - Variable definitions
    - Conditional sections
    - Loops and filters
    """
    
    name: str
    description: str
    format: ExportFormat
    template_content: str
    variables: list[TemplateVariable] = field(default_factory=list)
    css_content: str = ""  # For HTML templates
    
    def render(self, data: ReportData, context: dict[str, Any] | None = None) -> str:
        """Render template with data.
        
        Args:
            data: Report data to render
            context: Additional template context variables
            
        Returns:
            Rendered content string
        """
        if not HAS_JINJA2:
            raise ImportError("jinja2 required for template rendering")
        
        env = Environment(loader=BaseLoader())
        
        # Add custom filters
        env.filters["format_time"] = lambda x: f"{x:.2f}" if isinstance(x, float) else str(x)
        env.filters["format_percent"] = lambda x: f"{x*100:.1f}%" if isinstance(x, float) else str(x)
        env.filters["json_pretty"] = lambda x: json.dumps(x, indent=2, default=str)
        
        template = env.from_string(self.template_content)
        
        # Build context
        render_context = {
            "data": data,
            "title": data.title,
            "summary": data.summary,
            "raw_results": data.raw_results,
            "comparison": data.comparison,
            "insights": data.insights,
            "metadata": data.metadata,
            "generated_at": datetime.fromtimestamp(data.generated_at),
            "css": self.css_content,
        }
        
        if context:
            render_context.update(context)
        
        return template.render(**render_context)


class TemplateLibrary:
    """Library of pre-built export templates."""
    
    # Executive Summary Template
    EXECUTIVE_SUMMARY = ExportTemplate(
        name="executive_summary",
        description="High-level executive summary report",
        format=ExportFormat.HTML,
        template_content="""
<!DOCTYPE html>
<html>
<head>
    <title>{{ title }} - Executive Summary</title>
    <style>
        {{ css }}
        body { font-family: Arial, sans-serif; margin: 40px; }
        .header { background: #2c3e50; color: white; padding: 20px; border-radius: 8px; }
        .metric-card { display: inline-block; width: 200px; margin: 10px; padding: 20px; 
                       background: #ecf0f1; border-radius: 8px; text-align: center; }
        .metric-value { font-size: 32px; font-weight: bold; color: #3498db; }
        .metric-label { color: #7f8c8d; margin-top: 5px; }
        .section { margin: 20px 0; padding: 15px; border-left: 4px solid #3498db; }
        .insight { background: #e8f6f3; padding: 10px; margin: 5px 0; border-radius: 4px; }
        .recommendation { background: #fdebd0; padding: 15px; margin: 10px 0; border-radius: 8px; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ title }}</h1>
        <p>Generated: {{ generated_at.strftime('%Y-%m-%d %H:%M:%S') }}</p>
    </div>
    
    <h2>Key Metrics</h2>
    <div class="metrics">
        {% if summary.total_executions %}
        <div class="metric-card">
            <div class="metric-value">{{ summary.total_executions }}</div>
            <div class="metric-label">Total Executions</div>
        </div>
        {% endif %}
        {% if summary.backends_compared %}
        <div class="metric-card">
            <div class="metric-value">{{ summary.backends_compared }}</div>
            <div class="metric-label">Backends Compared</div>
        </div>
        {% endif %}
        {% if summary.average_execution_time %}
        <div class="metric-card">
            <div class="metric-value">{{ summary.average_execution_time | format_time }}ms</div>
            <div class="metric-label">Avg Execution Time</div>
        </div>
        {% endif %}
        {% if summary.result_agreement %}
        <div class="metric-card">
            <div class="metric-value">{{ summary.result_agreement | format_percent }}</div>
            <div class="metric-label">Result Agreement</div>
        </div>
        {% endif %}
    </div>
    
    {% if comparison and comparison.recommended_backend %}
    <div class="recommendation">
        <h3>Recommendation</h3>
        <p><strong>Recommended Backend:</strong> {{ comparison.recommended_backend }}</p>
        {% if comparison.recommendation_reason %}
        <p><strong>Reason:</strong> {{ comparison.recommendation_reason }}</p>
        {% endif %}
    </div>
    {% endif %}
    
    {% if insights %}
    <div class="section">
        <h2>Key Insights</h2>
        {% for insight in insights[:5] %}
        <div class="insight">{{ insight }}</div>
        {% endfor %}
    </div>
    {% endif %}
</body>
</html>
        """,
        css_content="",
    )
    
    # Technical Detail Template
    TECHNICAL_DETAIL = ExportTemplate(
        name="technical_detail",
        description="Detailed technical report with all data",
        format=ExportFormat.HTML,
        template_content="""
<!DOCTYPE html>
<html>
<head>
    <title>{{ title }} - Technical Report</title>
    <style>
        body { font-family: 'Courier New', monospace; margin: 40px; background: #1a1a2e; color: #eee; }
        .header { border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th { background: #16213e; color: #3498db; padding: 10px; text-align: left; }
        td { padding: 10px; border-bottom: 1px solid #333; }
        tr:hover { background: #16213e; }
        .code-block { background: #0f0f23; padding: 15px; border-radius: 4px; overflow-x: auto; }
        pre { margin: 0; color: #0f0; }
        .section-title { color: #3498db; border-left: 4px solid #3498db; padding-left: 10px; }
        .stat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; }
        .stat-box { background: #16213e; padding: 15px; border-radius: 4px; }
        .stat-value { font-size: 24px; color: #2ecc71; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ title }}</h1>
        <p>Generated: {{ generated_at.isoformat() }}</p>
    </div>
    
    {% if comparison and comparison.execution_times %}
    <h2 class="section-title">Execution Performance</h2>
    <table>
        <thead>
            <tr>
                <th>Backend</th>
                <th>Time (ms)</th>
                <th>Memory (MB)</th>
                <th>Ratio</th>
            </tr>
        </thead>
        <tbody>
            {% for backend, time in comparison.execution_times.items() %}
            <tr>
                <td>{{ backend }}</td>
                <td>{{ time | format_time }}</td>
                <td>{{ comparison.memory_peaks.get(backend, 0) | format_time }}</td>
                <td>{{ comparison.time_ratios.get(backend, 1.0) | format_time }}x</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
    {% endif %}
    
    {% if raw_results %}
    <h2 class="section-title">Raw Results</h2>
    <div class="code-block">
        <pre>{{ raw_results | json_pretty }}</pre>
    </div>
    {% endif %}
    
    {% if metadata %}
    <h2 class="section-title">Metadata</h2>
    <div class="code-block">
        <pre>{{ metadata | json_pretty }}</pre>
    </div>
    {% endif %}
</body>
</html>
        """,
    )
    
    # Comparison Matrix Template
    COMPARISON_MATRIX = ExportTemplate(
        name="comparison_matrix",
        description="Backend comparison matrix view",
        format=ExportFormat.MARKDOWN,
        template_content="""
# {{ title }}

Generated: {{ generated_at.strftime('%Y-%m-%d %H:%M:%S') }}

## Backend Comparison Matrix

{% if comparison and comparison.execution_times %}
| Metric | {% for backend in comparison.execution_times.keys() %}{{ backend }} | {% endfor %}
|--------|{% for _ in comparison.execution_times %}---------|{% endfor %}
| Time (ms) | {% for backend, time in comparison.execution_times.items() %}{{ "%.2f"|format(time) }} | {% endfor %}
| Memory (MB) | {% for backend in comparison.execution_times.keys() %}{{ "%.2f"|format(comparison.memory_peaks.get(backend, 0)) }} | {% endfor %}
| Time Ratio | {% for backend in comparison.execution_times.keys() %}{{ "%.2f"|format(comparison.time_ratios.get(backend, 1.0)) }}x | {% endfor %}
{% endif %}

{% if comparison and comparison.pairwise_agreements %}
## Pairwise Agreement Matrix

| Backend | {% for b in comparison.pairwise_agreements.keys() %}{{ b }} | {% endfor %}
|---------|{% for _ in comparison.pairwise_agreements %}---------|{% endfor %}
{% for b1, agreements in comparison.pairwise_agreements.items() %}
| {{ b1 }} | {% for b2 in comparison.pairwise_agreements.keys() %}{{ "%.1f%%"|format(agreements.get(b2, 0) * 100) }} | {% endfor %}
{% endfor %}
{% endif %}

{% if comparison and comparison.recommended_backend %}
## Recommendation

**Recommended Backend:** {{ comparison.recommended_backend }}

{% if comparison.recommendation_reason %}
**Reason:** {{ comparison.recommendation_reason }}
{% endif %}
{% endif %}

{% if insights %}
## Key Insights

{% for insight in insights %}
- {{ insight }}
{% endfor %}
{% endif %}
        """,
    )
    
    @classmethod
    def get_template(cls, name: str) -> ExportTemplate | None:
        """Get template by name."""
        templates = {
            "executive_summary": cls.EXECUTIVE_SUMMARY,
            "technical_detail": cls.TECHNICAL_DETAIL,
            "comparison_matrix": cls.COMPARISON_MATRIX,
        }
        return templates.get(name)
    
    @classmethod
    def list_templates(cls) -> list[str]:
        """List available template names."""
        return ["executive_summary", "technical_detail", "comparison_matrix"]


class TemplateExporter(BaseExporter):
    """Exporter that uses custom templates."""
    
    def __init__(self, template: ExportTemplate) -> None:
        """Initialize with template."""
        self._template = template
    
    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export using template."""
        import time
        start_time = time.perf_counter()
        
        try:
            content = self._template.render(data)
            
            if options.stream_output:
                return ExportResult(
                    success=True,
                    format=self._template.format,
                    output_path=None,
                    file_size_bytes=len(content.encode("utf-8")),
                    export_time_ms=(time.perf_counter() - start_time) * 1000,
                    content=content,
                )
            
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=self._template.format,
                    output_path=None,
                    error="Output path required",
                )
            
            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(content, encoding="utf-8")
            
            return ExportResult(
                success=True,
                format=self._template.format,
                output_path=options.output_path,
                file_size_bytes=options.output_path.stat().st_size,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                format=self._template.format,
                output_path=options.output_path,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
                error=str(e),
            )


# =============================================================================
# Data Visualization Integration (Feature - Export Engine)
# =============================================================================


@dataclass
class ChartConfig:
    """Configuration for chart generation."""
    
    type: str  # 'bar', 'line', 'pie', 'scatter', 'heatmap'
    title: str = ""
    x_label: str = ""
    y_label: str = ""
    width: int = 800
    height: int = 500
    colors: list[str] | None = None
    show_legend: bool = True
    show_grid: bool = True


class VisualizationGenerator:
    """Generate visualizations for export reports.
    
    Supports:
    - ASCII charts (no dependencies)
    - SVG charts (built-in)
    - Matplotlib integration (optional)
    - Chart embedding in HTML/Markdown exports
    """
    
    DEFAULT_COLORS = [
        "#3498db", "#2ecc71", "#e74c3c", "#f39c12", 
        "#9b59b6", "#1abc9c", "#34495e", "#e67e22",
    ]
    
    def __init__(self) -> None:
        """Initialize visualization generator."""
        self._has_matplotlib = False
        try:
            import matplotlib
            matplotlib.use("Agg")  # Non-interactive backend
            self._has_matplotlib = True
        except ImportError:
            pass
    
    def generate_ascii_bar_chart(
        self,
        data: dict[str, float],
        title: str = "",
        width: int = 50,
    ) -> str:
        """Generate ASCII bar chart.
        
        Args:
            data: Dictionary of label -> value
            title: Chart title
            width: Maximum bar width in characters
            
        Returns:
            ASCII chart string
        """
        if not data:
            return "No data to visualize"
        
        lines = []
        if title:
            lines.append(f"  {title}")
            lines.append("  " + "=" * len(title))
        
        max_value = max(data.values())
        max_label_len = max(len(str(k)) for k in data.keys())
        
        for label, value in data.items():
            bar_len = int((value / max_value) * width) if max_value > 0 else 0
            bar = "█" * bar_len
            label_padded = str(label).ljust(max_label_len)
            lines.append(f"  {label_padded} │{bar} {value:.2f}")
        
        return "\n".join(lines)
    
    def generate_ascii_horizontal_bar(
        self,
        data: dict[str, float],
        title: str = "",
    ) -> str:
        """Generate horizontal ASCII bar chart with proportional bars."""
        return self.generate_ascii_bar_chart(data, title)
    
    def generate_svg_bar_chart(
        self,
        data: dict[str, float],
        config: ChartConfig | None = None,
    ) -> str:
        """Generate SVG bar chart.
        
        Args:
            data: Dictionary of label -> value
            config: Chart configuration
            
        Returns:
            SVG string
        """
        config = config or ChartConfig(type="bar")
        colors = config.colors or self.DEFAULT_COLORS
        
        if not data:
            return '<svg><text x="10" y="20">No data</text></svg>'
        
        width = config.width
        height = config.height
        margin = 60
        bar_width = (width - 2 * margin) / len(data)
        max_value = max(data.values()) if data.values() else 1
        
        svg_parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}">',
            f'<rect width="100%" height="100%" fill="white"/>',
        ]
        
        # Title
        if config.title:
            svg_parts.append(
                f'<text x="{width/2}" y="25" text-anchor="middle" '
                f'font-size="16" font-weight="bold">{config.title}</text>'
            )
        
        # Bars
        for i, (label, value) in enumerate(data.items()):
            x = margin + i * bar_width + bar_width * 0.1
            bar_h = (value / max_value) * (height - 2 * margin - 30) if max_value > 0 else 0
            y = height - margin - bar_h
            color = colors[i % len(colors)]
            
            svg_parts.append(
                f'<rect x="{x}" y="{y}" width="{bar_width * 0.8}" height="{bar_h}" '
                f'fill="{color}" rx="2"/>'
            )
            
            # Label
            label_x = x + bar_width * 0.4
            svg_parts.append(
                f'<text x="{label_x}" y="{height - margin + 15}" '
                f'text-anchor="middle" font-size="10">{label[:10]}</text>'
            )
            
            # Value
            svg_parts.append(
                f'<text x="{label_x}" y="{y - 5}" '
                f'text-anchor="middle" font-size="10">{value:.1f}</text>'
            )
        
        svg_parts.append('</svg>')
        return "\n".join(svg_parts)
    
    def generate_svg_line_chart(
        self,
        data: dict[str, list[float]],
        x_labels: list[str] | None = None,
        config: ChartConfig | None = None,
    ) -> str:
        """Generate SVG line chart.
        
        Args:
            data: Dictionary of series_name -> values
            x_labels: Labels for x-axis
            config: Chart configuration
            
        Returns:
            SVG string
        """
        config = config or ChartConfig(type="line")
        colors = config.colors or self.DEFAULT_COLORS
        
        if not data:
            return '<svg><text x="10" y="20">No data</text></svg>'
        
        width = config.width
        height = config.height
        margin = 60
        
        # Find data bounds
        all_values = [v for values in data.values() for v in values]
        if not all_values:
            return '<svg><text x="10" y="20">No data</text></svg>'
        
        max_value = max(all_values)
        min_value = min(all_values)
        value_range = max_value - min_value or 1
        
        num_points = max(len(v) for v in data.values())
        
        svg_parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}">',
            '<rect width="100%" height="100%" fill="white"/>',
        ]
        
        # Title
        if config.title:
            svg_parts.append(
                f'<text x="{width/2}" y="25" text-anchor="middle" '
                f'font-size="16" font-weight="bold">{config.title}</text>'
            )
        
        # Grid lines
        if config.show_grid:
            for i in range(5):
                y = margin + i * (height - 2 * margin) / 4
                svg_parts.append(
                    f'<line x1="{margin}" y1="{y}" x2="{width - margin}" y2="{y}" '
                    f'stroke="#ddd" stroke-dasharray="5,5"/>'
                )
        
        # Lines for each series
        for idx, (name, values) in enumerate(data.items()):
            color = colors[idx % len(colors)]
            points = []
            
            for i, value in enumerate(values):
                x = margin + i * (width - 2 * margin) / (num_points - 1 or 1)
                y = height - margin - ((value - min_value) / value_range) * (height - 2 * margin - 30)
                points.append(f"{x},{y}")
            
            if points:
                svg_parts.append(
                    f'<polyline points="{" ".join(points)}" '
                    f'fill="none" stroke="{color}" stroke-width="2"/>'
                )
        
        # Legend
        if config.show_legend and len(data) > 1:
            legend_y = height - 20
            for idx, name in enumerate(data.keys()):
                color = colors[idx % len(colors)]
                x = margin + idx * 100
                svg_parts.append(
                    f'<rect x="{x}" y="{legend_y}" width="15" height="10" fill="{color}"/>'
                )
                svg_parts.append(
                    f'<text x="{x + 20}" y="{legend_y + 9}" font-size="10">{name}</text>'
                )
        
        svg_parts.append('</svg>')
        return "\n".join(svg_parts)
    
    def generate_svg_pie_chart(
        self,
        data: dict[str, float],
        config: ChartConfig | None = None,
    ) -> str:
        """Generate SVG pie chart.
        
        Args:
            data: Dictionary of label -> value
            config: Chart configuration
            
        Returns:
            SVG string
        """
        config = config or ChartConfig(type="pie")
        colors = config.colors or self.DEFAULT_COLORS
        
        if not data:
            return '<svg><text x="10" y="20">No data</text></svg>'
        
        width = config.width
        height = config.height
        cx, cy = width / 2, height / 2
        radius = min(width, height) / 2 - 60
        
        total = sum(data.values())
        if total == 0:
            return '<svg><text x="10" y="20">No data</text></svg>'
        
        svg_parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}">',
            '<rect width="100%" height="100%" fill="white"/>',
        ]
        
        # Title
        if config.title:
            svg_parts.append(
                f'<text x="{width/2}" y="25" text-anchor="middle" '
                f'font-size="16" font-weight="bold">{config.title}</text>'
            )
        
        # Pie slices
        import math
        start_angle = 0
        
        for idx, (label, value) in enumerate(data.items()):
            if value <= 0:
                continue
                
            slice_angle = (value / total) * 2 * math.pi
            end_angle = start_angle + slice_angle
            
            # Calculate arc path
            x1 = cx + radius * math.cos(start_angle)
            y1 = cy + radius * math.sin(start_angle)
            x2 = cx + radius * math.cos(end_angle)
            y2 = cy + radius * math.sin(end_angle)
            
            large_arc = 1 if slice_angle > math.pi else 0
            color = colors[idx % len(colors)]
            
            path = f"M {cx} {cy} L {x1} {y1} A {radius} {radius} 0 {large_arc} 1 {x2} {y2} Z"
            svg_parts.append(f'<path d="{path}" fill="{color}" stroke="white" stroke-width="2"/>')
            
            # Label
            mid_angle = start_angle + slice_angle / 2
            label_x = cx + (radius + 20) * math.cos(mid_angle)
            label_y = cy + (radius + 20) * math.sin(mid_angle)
            pct = value / total * 100
            svg_parts.append(
                f'<text x="{label_x}" y="{label_y}" text-anchor="middle" '
                f'font-size="10">{label}: {pct:.1f}%</text>'
            )
            
            start_angle = end_angle
        
        svg_parts.append('</svg>')
        return "\n".join(svg_parts)
    
    def embed_chart_in_html(self, svg_content: str, caption: str = "") -> str:
        """Embed SVG chart in HTML figure element."""
        return f'''
<figure style="text-align: center; margin: 20px 0;">
    {svg_content}
    {f'<figcaption style="color: #666; margin-top: 10px;">{caption}</figcaption>' if caption else ''}
</figure>
        '''


# =============================================================================
# Custom Format Support (Feature - Export Engine)
# =============================================================================


class CustomFormat(Enum):
    """Custom export formats beyond standard ones."""
    
    LATEX = "latex"
    QASM = "qasm"
    JUPYTER = "ipynb"
    PDF = "pdf"


@dataclass
class CustomFormatConfig:
    """Configuration for custom format export."""
    
    format: CustomFormat
    template: str | None = None
    options: dict[str, Any] = field(default_factory=dict)


class LaTeXExporter(BaseExporter):
    """Export data to LaTeX format for academic papers."""
    
    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data to LaTeX format."""
        import time
        start_time = time.perf_counter()
        
        try:
            content = self._generate_latex(data, options)
            
            if options.stream_output:
                return ExportResult(
                    success=True,
                    format=ExportFormat.MARKDOWN,  # Using MD as placeholder
                    output_path=None,
                    file_size_bytes=len(content.encode("utf-8")),
                    export_time_ms=(time.perf_counter() - start_time) * 1000,
                    content=content,
                )
            
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.MARKDOWN,
                    output_path=None,
                    error="Output path required",
                )
            
            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(content, encoding="utf-8")
            
            return ExportResult(
                success=True,
                format=ExportFormat.MARKDOWN,
                output_path=options.output_path,
                file_size_bytes=options.output_path.stat().st_size,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.MARKDOWN,
                output_path=options.output_path,
                error=str(e),
            )
    
    def _generate_latex(self, data: ReportData, options: ExportOptions) -> str:
        """Generate LaTeX document."""
        lines = [
            r"\documentclass{article}",
            r"\usepackage{booktabs}",
            r"\usepackage{graphicx}",
            r"\usepackage{hyperref}",
            r"\usepackage{amsmath}",
            "",
            r"\title{" + self._escape_latex(data.title) + "}",
            r"\date{\today}",
            "",
            r"\begin{document}",
            r"\maketitle",
            "",
        ]
        
        # Abstract/Summary
        if data.summary:
            lines.append(r"\begin{abstract}")
            lines.append(self._format_summary_latex(data.summary))
            lines.append(r"\end{abstract}")
            lines.append("")
        
        # Results table
        if data.comparison and data.comparison.get("execution_times"):
            lines.append(r"\section{Backend Comparison}")
            lines.append(self._format_comparison_table_latex(data.comparison))
            lines.append("")
        
        # Insights
        if data.insights:
            lines.append(r"\section{Key Insights}")
            lines.append(r"\begin{itemize}")
            for insight in data.insights[:10]:
                lines.append(r"  \item " + self._escape_latex(insight))
            lines.append(r"\end{itemize}")
            lines.append("")
        
        lines.append(r"\end{document}")
        
        return "\n".join(lines)
    
    def _escape_latex(self, text: str) -> str:
        """Escape special LaTeX characters."""
        replacements = [
            ("\\", r"\textbackslash{}"),
            ("{", r"\{"),
            ("}", r"\}"),
            ("&", r"\&"),
            ("%", r"\%"),
            ("$", r"\$"),
            ("#", r"\#"),
            ("_", r"\_"),
            ("^", r"\^{}"),
            ("~", r"\~{}"),
        ]
        for old, new in replacements:
            text = text.replace(old, new)
        return text
    
    def _format_summary_latex(self, summary: dict[str, Any]) -> str:
        """Format summary for LaTeX."""
        parts = []
        for key, value in summary.items():
            if isinstance(value, float):
                parts.append(f"{key}: {value:.2f}")
            else:
                parts.append(f"{key}: {value}")
        return ". ".join(parts) + "."
    
    def _format_comparison_table_latex(self, comparison: dict[str, Any]) -> str:
        """Format comparison data as LaTeX table."""
        execution_times = comparison.get("execution_times", {})
        if not execution_times:
            return ""
        
        backends = list(execution_times.keys())
        lines = [
            r"\begin{table}[h]",
            r"\centering",
            r"\begin{tabular}{l" + "r" * len(backends) + "}",
            r"\toprule",
            "Metric & " + " & ".join(backends) + r" \\",
            r"\midrule",
        ]
        
        # Execution times
        times_row = "Time (ms) & " + " & ".join(
            f"{execution_times[b]:.2f}" for b in backends
        ) + r" \\"
        lines.append(times_row)
        
        # Memory usage
        if comparison.get("memory_peaks"):
            mem_row = "Memory (MB) & " + " & ".join(
                f"{comparison['memory_peaks'].get(b, 0):.2f}" for b in backends
            ) + r" \\"
            lines.append(mem_row)
        
        lines.extend([
            r"\bottomrule",
            r"\end{tabular}",
            r"\caption{Backend Performance Comparison}",
            r"\label{tab:comparison}",
            r"\end{table}",
        ])
        
        return "\n".join(lines)


class JupyterNotebookExporter(BaseExporter):
    """Export data to Jupyter Notebook format."""
    
    def export(self, data: ReportData, options: ExportOptions) -> ExportResult:
        """Export data as Jupyter notebook."""
        import time
        start_time = time.perf_counter()
        
        try:
            notebook = self._generate_notebook(data, options)
            content = json.dumps(notebook, indent=2)
            
            if options.stream_output:
                return ExportResult(
                    success=True,
                    format=ExportFormat.JSON,
                    output_path=None,
                    file_size_bytes=len(content.encode("utf-8")),
                    export_time_ms=(time.perf_counter() - start_time) * 1000,
                    content=content,
                )
            
            if not options.output_path:
                return ExportResult(
                    success=False,
                    format=ExportFormat.JSON,
                    output_path=None,
                    error="Output path required",
                )
            
            options.output_path.parent.mkdir(parents=True, exist_ok=True)
            options.output_path.write_text(content, encoding="utf-8")
            
            return ExportResult(
                success=True,
                format=ExportFormat.JSON,
                output_path=options.output_path,
                file_size_bytes=options.output_path.stat().st_size,
                export_time_ms=(time.perf_counter() - start_time) * 1000,
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                format=ExportFormat.JSON,
                output_path=options.output_path,
                error=str(e),
            )
    
    def _generate_notebook(self, data: ReportData, options: ExportOptions) -> dict:
        """Generate Jupyter notebook structure."""
        cells = []
        
        # Title cell
        cells.append(self._markdown_cell(f"# {data.title}\n\nGenerated: {datetime.fromtimestamp(data.generated_at).isoformat()}"))
        
        # Summary cell
        if data.summary:
            cells.append(self._markdown_cell("## Summary"))
            cells.append(self._code_cell(f"summary = {json.dumps(data.summary, indent=2)}\nprint(summary)"))
        
        # Comparison visualization
        if data.comparison and data.comparison.get("execution_times"):
            cells.append(self._markdown_cell("## Backend Comparison"))
            
            viz_code = """
import matplotlib.pyplot as plt

execution_times = """ + json.dumps(data.comparison.get("execution_times", {})) + """

plt.figure(figsize=(10, 6))
plt.bar(execution_times.keys(), execution_times.values(), color='steelblue')
plt.xlabel('Backend')
plt.ylabel('Execution Time (ms)')
plt.title('Backend Execution Times')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()
"""
            cells.append(self._code_cell(viz_code))
        
        # Raw results
        if data.raw_results:
            cells.append(self._markdown_cell("## Raw Results"))
            cells.append(self._code_cell(f"results = {json.dumps(data.raw_results[:5], indent=2)}"))
        
        # Insights
        if data.insights:
            cells.append(self._markdown_cell("## Insights\n\n" + "\n".join(f"- {i}" for i in data.insights)))
        
        return {
            "nbformat": 4,
            "nbformat_minor": 5,
            "metadata": {
                "kernelspec": {
                    "display_name": "Python 3",
                    "language": "python",
                    "name": "python3",
                },
                "language_info": {
                    "name": "python",
                    "version": "3.10.0",
                },
            },
            "cells": cells,
        }
    
    def _markdown_cell(self, content: str) -> dict:
        """Create markdown cell."""
        return {
            "cell_type": "markdown",
            "metadata": {},
            "source": content.split("\n"),
        }
    
    def _code_cell(self, content: str) -> dict:
        """Create code cell."""
        return {
            "cell_type": "code",
            "metadata": {},
            "source": content.split("\n"),
            "outputs": [],
            "execution_count": None,
        }


class EnhancedExportEngine(ExportEngine):
    """Extended export engine with advanced features.
    
    Adds:
    - Template-based exports
    - Visualization integration
    - Custom format support
    - Batch export operations
    """
    
    def __init__(self) -> None:
        """Initialize enhanced export engine."""
        super().__init__()
        self._templates: dict[str, ExportTemplate] = {}
        self._visualizer = VisualizationGenerator()
        self._custom_exporters: dict[CustomFormat, BaseExporter] = {
            CustomFormat.LATEX: LaTeXExporter(),
            CustomFormat.JUPYTER: JupyterNotebookExporter(),
        }
        
        # Load built-in templates
        for name in TemplateLibrary.list_templates():
            template = TemplateLibrary.get_template(name)
            if template:
                self._templates[name] = template
    
    def register_template(self, template: ExportTemplate) -> None:
        """Register a custom template."""
        self._templates[template.name] = template
    
    def export_with_template(
        self,
        data: ReportData,
        template_name: str,
        output_path: Path | None = None,
        context: dict[str, Any] | None = None,
    ) -> ExportResult:
        """Export using a named template.
        
        Args:
            data: Report data to export
            template_name: Name of template to use
            output_path: Output file path
            context: Additional template context
            
        Returns:
            Export result
        """
        template = self._templates.get(template_name)
        if not template:
            return ExportResult(
                success=False,
                format=ExportFormat.HTML,
                output_path=output_path,
                error=f"Template not found: {template_name}",
            )
        
        exporter = TemplateExporter(template)
        options = ExportOptions(
            format=template.format,
            output_path=output_path,
            stream_output=output_path is None,
        )
        
        return exporter.export(data, options)
    
    def export_custom_format(
        self,
        data: ReportData,
        format: CustomFormat,
        output_path: Path | None = None,
        **kwargs: Any,
    ) -> ExportResult:
        """Export to a custom format.
        
        Args:
            data: Report data to export
            format: Custom format (LATEX, JUPYTER, etc.)
            output_path: Output file path
            **kwargs: Additional options
            
        Returns:
            Export result
        """
        exporter = self._custom_exporters.get(format)
        if not exporter:
            return ExportResult(
                success=False,
                format=ExportFormat.JSON,
                output_path=output_path,
                error=f"No exporter for custom format: {format.value}",
            )
        
        options = ExportOptions(
            format=ExportFormat.JSON,
            output_path=output_path,
            stream_output=output_path is None,
            **kwargs,
        )
        
        return exporter.export(data, options)
    
    def export_with_visualizations(
        self,
        data: ReportData,
        format: ExportFormat = ExportFormat.HTML,
        output_path: Path | None = None,
        charts: list[str] | None = None,
    ) -> ExportResult:
        """Export with embedded visualizations.
        
        Args:
            data: Report data to export
            format: Export format (HTML recommended)
            output_path: Output file path
            charts: List of chart types to include ('execution_times', 'memory', 'agreement')
            
        Returns:
            Export result with embedded charts
        """
        charts = charts or ["execution_times"]
        
        # Generate visualizations
        viz_content = []
        
        if "execution_times" in charts and data.comparison:
            exec_times = data.comparison.get("execution_times", {})
            if exec_times:
                svg = self._visualizer.generate_svg_bar_chart(
                    exec_times,
                    ChartConfig(type="bar", title="Execution Times by Backend"),
                )
                viz_content.append(self._visualizer.embed_chart_in_html(svg, "Backend Execution Times"))
        
        if "memory" in charts and data.comparison:
            memory = data.comparison.get("memory_peaks", {})
            if memory:
                svg = self._visualizer.generate_svg_bar_chart(
                    memory,
                    ChartConfig(type="bar", title="Memory Usage by Backend"),
                )
                viz_content.append(self._visualizer.embed_chart_in_html(svg, "Backend Memory Usage"))
        
        # Add visualizations to custom sections
        enhanced_data = ReportData(
            title=data.title,
            generated_at=data.generated_at,
            summary=data.summary,
            raw_results=data.raw_results,
            comparison=data.comparison,
            insights=data.insights,
            metadata=data.metadata,
            custom_sections={
                **data.custom_sections,
                "visualizations_html": "\n".join(viz_content),
            },
        )
        
        return self.export(enhanced_data, format, output_path)
    
    def list_templates(self) -> list[str]:
        """List available template names."""
        return list(self._templates.keys())
    
    def list_custom_formats(self) -> list[str]:
        """List available custom formats."""
        return [f.value for f in self._custom_exporters.keys()]
